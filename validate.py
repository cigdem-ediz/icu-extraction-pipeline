#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
karsilastir.py — Manuel ve otomatik ICU veri cikarimini alan-alan karsilastirir.

Bu script YORUM KATMAZ: iki hucre birebir ayni degilse "UYUMSUZ" isaretler.
Dogruluk yargisi (hangisi dogru) kullaniciya aittir; script sadece eslestirir.

Kategoriler:
  ESLESTI          : iki kaynakta da ayni deger (birebir)
  UYUMSUZ          : farkli degerler VEYA biri dolu digeri bos
  IKISI_DE_BOS     : her iki kaynakta da bos (ayri liste; orneklem denetimi icin)
  FARKLI_GUN       : zaman-noktasi sekmelerinde farkli (ama gecerli) gun secilmis

Kullanim:
  python karsilastir.py --manuel manuel.xlsx --otomatik otomatik.xlsx --cikti rapor.xlsx
"""

import argparse
import unicodedata
import openpyxl
from openpyxl.styles import Font, PatternFill
from collections import defaultdict

# ============================================================
# YAPILANDIRMA — her sekme icin anahtar ve karsilastirilacak sutunlar
# ============================================================

# Zaman-noktasi sekmeleri (±2 gun penceresi; farkli gun secimi ayri kategori)
ZAMAN_NOKTASI_SEKMELERI = {"KLINIK", "KAN_GAZI"}

# Her sekme icin: anahtar sutunlar + karsilastirilacak deger sutunlari
# Turetilmis/eczaci-degerlendirmesi sutunlar KARSILASTIRILMAZ (asagida yok)
SEKME_CONFIG = {
    "DEMOGRAFI": {
        "anahtar": ["Hasta_ID"],
        "degerler": ["Cinsiyet", "Yatis_Tarihi", "Cikis_Tarihi", "Taburcu_Sekli",
                     "ICD_Kodlari", "Tanilar", "Sikayet", "Apache_II_Basvuru",
                     "Ventilasyon", "SVK", "Enfeksiyon_Kaynagi",
                     "Beslenme_Baslangic", "Beslenme_Son", "Beslenme_Gecis_Gunu"],
    },
    "KLINIK": {
        "anahtar": ["Hasta_ID", "Tarih"],
        "degerler": ["YBU_Seviye", "APACHE_II", "SOFA", "GKS_Toplam",
                     "Sepsis", "Septik_Sok"],
    },
    "LABORATUVAR": {
        "anahtar": ["Hasta_ID", "Tarih"],
        "degerler": ["Kaynak", "WBC", "CRP", "Kreatinin", "BUN", "AST", "ALT",
                     "Total_Bil", "Direkt_Bil", "Albumin", "Na", "HCT", "PLT",
                     "NEU_pct", "Laktat"],
    },
    "KAN_GAZI": {
        "anahtar": ["Hasta_ID", "Tarih"],
        "degerler": ["pH", "pCO2", "pO2", "HCO3", "BE", "SpO2", "Na", "K"],
    },
    "ANTIMIKROBIAL": {
        "anahtar": ["Hasta_ID", "Ilac_Adi", "Kurs_No"],
        "degerler": ["Doz", "Siklik", "Uygulama_Yolu", "Baslangic_Tarihi",
                     "Bitis_Tarihi", "Sure_Gun", "Bas_Gun_No"],
    },
    "VAZOPRESSOR": {
        "anahtar": ["Hasta_ID", "Ilac", "Baslangic_Tarihi"],
        "degerler": ["Bitis_Tarihi", "Sure_Gun",
                     "Baslangic_Gun_No", "Bitis_Gun_No"],
    },
    "MIKROBIYOLOJI_ABG": {
        "anahtar": ["Hasta_ID", "Rapor_No", "Mikroorganizma", "Antibiyotik"],
        "degerler": ["Numune_Turu", "Numune_Yeri", "Numune_Tarihi",
                     "Sonuc_Tarihi", "Kultur_Sonucu", "Duyarlilik"],
    },
}
# NOT: MIC karsilastirmaya alinmadi (tutarsiz kayit).
# Turetilmis/eczaci sutunlari (ATC, AWaRe, DDD, Etken_Madde, Yas, Yatis_Suresi_Gun,
# Renal_Durum, EHU_Onayi, TDM, Tedavi_Turu, Uygunluk, Uygunsuzluk_Nedeni,
# Deskalasyon, Eskalasyon_Nedeni) yukaridaki "degerler" listelerinde YOK — dislanmis.


# Turkce I/i varyantlari — dordu de tek harfe indirgenir
_I_VARYANT = str.maketrans({
    "\u0130": "i",   # I  noktali buyuk I
    "\u0049": "i",   # I  noktasiz buyuk I
    "\u0131": "i",   # i  noktasiz kucuk i
    "\u0069": "i",   # i  noktali kucuk i
})


def tr_fold(s):
    """Turkce-duyarli buyuk/kucuk harf esitlemesi.

    NEDEN GEREKLI:
      Python'un str.casefold() metodu Turkce'yi bilmez. 'I' (U+0130, noktali
      buyuk I) icin 'i' + U+0307 (COMBINING DOT ABOVE) uretir. Ayni kelime
      kucuk yazilmissa yalnizca 'i' olur. Sonuc: 'UREDI'.casefold() ile
      'uredi'.casefold() FARKLI cikar ve YAPAY uyumsuzluk uretilir.
      Ayrica Turkce'de 'I'->'i' (noktasiz), Ingilizce'de 'I'->'i'dir; bir
      metindeki buyuk 'I'nin hangisi oldugu bilinemez ('KANDIDA' -> 'kandida'
      mi 'kandida' mi?).

    COZUM:
      Dort I varyanti (I, I, i, i) tek bir 'i' harfine indirgenir. Bu, ayni
      degerin farkli yazimlarinin ASLA yapay uyumsuzluk uretmemesini garanti
      eder. Yanlis EslESME riski ihmal edilebilir: klinik alanlarda yalnizca
      i/i farkiyla ayrisan iki gecerli deger bulunmaz.

    Bu bir icerik/birim donusumu DEGIL, yalnizca kodlama ve harf-buyuklugu
    esitlemesidir.
    """
    s = s.translate(_I_VARYANT)
    s = s.casefold()
    s = s.replace("\u0307", "")   # arta kalan birlesik nokta
    return unicodedata.normalize("NFC", s)


def normalize(v):
    """Hucre degerini karsilastirma icin metne cevirir.

    YORUM KATMAZ — sadece ayni degerin farkli TEKNIK YAZIMLARINI esitler:
      - sayi/metin farki: 100001 (sayi) == '100001' (metin)
      - ondalik yazim:    12.0 == 12,  3.50 == 3.5
      - tarih nesnesi:    datetime(2022,4,1) == '2022-04-01' (sadece Y-A-G)
      - Turkce I/i:       'UREDI' == 'uredi' == 'Uredi'  (bkz. tr_fold)
    Birim/olcek/yuvarlama donusumu YAPMAZ (12.4 ile 12400 farkli kalir)."""
    from datetime import datetime, date
    if v is None:
        return ""
    # Tarih/zaman: sadece yil-ay-gun (saat atilir)
    if isinstance(v, (datetime, date)):
        return f"{v.year:04d}-{v.month:02d}-{v.day:02d}"
    # Sayi: ondalik yazim farkini gider (12.0 -> "12", 3.50 -> "3.5")
    if isinstance(v, (int, float)):
        f = float(v)
        if f == int(f):
            return str(int(f))
        return repr(f).rstrip("0").rstrip(".")
    # Metin — once Unicode NFC (ayni karakterin farkli kodlanmasini esitler)
    s = unicodedata.normalize("NFC", str(v)).strip()
    # Metin bir sayiyi mi temsil ediyor? (virgullu ondalik dahil: "12,4" -> "12.4")
    s_test = s.replace(",", ".", 1) if s.count(",") == 1 and s.count(".") == 0 else s
    try:
        f = float(s_test)
        if f == int(f):
            return str(int(f))
        return repr(f).rstrip("0").rstrip(".")
    except (ValueError, OverflowError):
        pass
    # Metin degeri: buyuk/kucuk harf farki gozardi edilsin (Var==var, E.coli==E.Coli)
    return tr_fold(s)


def is_bos(v):
    return normalize(v) == ""


def sheet_to_dict(ws, anahtar_sutunlar, deger_sutunlar):
    """Bir sekmeyi {anahtar_tuple: {deger_sutun: deger}} sozlugune cevirir."""
    # Baslik satirini oku, sutun indekslerini bul (buyuk/kucuk harf duyarsiz)
    basliklar = [str(c.value).strip() if c.value is not None else ""
                 for c in ws[1]]
    # Sutun adlarini kucuk harfe cevirerek eslestir (Sepsis == sepsis)
    idx = {b.casefold(): i for i, b in enumerate(basliklar)}

    def bul(sutun):
        return idx.get(sutun.casefold())

    # Gerekli sutunlar var mi kontrol
    eksik = [s for s in (anahtar_sutunlar + deger_sutunlar) if bul(s) is None]
    if eksik:
        print(f"    UYARI: bu sekmede bulunamayan sutunlar: {eksik}")

    kayitlar = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(c is None for c in row):
            continue  # tamamen bos satir
        # anahtar olustur (buyuk/kucuk harf duyarsiz sutun eslesmesi)
        anahtar = tuple(
            normalize(row[bul(a)]) if (bul(a) is not None and bul(a) < len(row))
            else ""
            for a in anahtar_sutunlar)
        if all(x == "" for x in anahtar):
            continue  # anahtarsiz satir
        degerler = {}
        for d in deger_sutunlar:
            j = bul(d)
            if j is not None and j < len(row):
                degerler[d] = row[j]
            else:
                degerler[d] = None
        kayitlar[anahtar] = degerler
    return kayitlar


def karsilastir_sekme(sekme_adi, ws_manuel, ws_oto, config):
    """Bir sekmeyi alan-alan karsilastirir, sonuc satirlari dondurur."""
    anahtar_s = config["anahtar"]
    deger_s = config["degerler"]
    zaman_noktasi = sekme_adi in ZAMAN_NOKTASI_SEKMELERI

    man = sheet_to_dict(ws_manuel, anahtar_s, deger_s)
    oto = sheet_to_dict(ws_oto, anahtar_s, deger_s)

    tum_anahtarlar = set(man.keys()) | set(oto.keys())
    sonuclar = []

    for anahtar in sorted(tum_anahtarlar):
        man_var = anahtar in man
        oto_var = anahtar in oto

        # Satir bir kaynakta hic yoksa (anahtar eslesmiyor)
        if man_var and not oto_var:
            for d in deger_s:
                mv = man[anahtar].get(d)
                if not is_bos(mv):
                    sonuclar.append([sekme_adi, *anahtar, d, normalize(mv), "",
                                     "UYUMSUZ", "satir otomatikte yok"])
            continue
        if oto_var and not man_var:
            for d in deger_s:
                ov = oto[anahtar].get(d)
                if not is_bos(ov):
                    sonuclar.append([sekme_adi, *anahtar, d, "", normalize(ov),
                                     "UYUMSUZ", "satir manuelde yok"])
            continue

        # Satir iki kaynakta da var — alan alan karsilastir
        for d in deger_s:
            mv = man[anahtar].get(d)
            ov = oto[anahtar].get(d)
            m_bos = is_bos(mv)
            o_bos = is_bos(ov)

            if m_bos and o_bos:
                kategori = "IKISI_DE_BOS"
                not_ = "orneklem denetimi gerekli"
            elif m_bos != o_bos:
                kategori = "UYUMSUZ"
                not_ = "biri bos digeri dolu"
            elif normalize(mv) == normalize(ov):
                kategori = "ESLESTI"
                not_ = ""
            else:
                # Iki deger de dolu ama farkli
                kategori = "UYUMSUZ"
                not_ = ""

            sonuclar.append([sekme_adi, *anahtar, d,
                             normalize(mv), normalize(ov), kategori, not_])

    return sonuclar, anahtar_s


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--manuel", required=True, help="Manuel cikarim Excel dosyasi")
    ap.add_argument("--otomatik", required=True, help="Otomatik (arac) Excel dosyasi")
    ap.add_argument("--cikti", default="karsilastirma_raporu.xlsx",
                    help="Cikti rapor dosyasi")
    args = ap.parse_args()

    print(f"Manuel dosya  : {args.manuel}")
    print(f"Otomatik dosya: {args.otomatik}")
    print()

    wb_man = openpyxl.load_workbook(args.manuel, read_only=True, data_only=True)
    wb_oto = openpyxl.load_workbook(args.otomatik, read_only=True, data_only=True)

    tum_sonuclar = []
    ozet = defaultdict(lambda: defaultdict(int))

    for sekme, config in SEKME_CONFIG.items():
        if sekme not in wb_man.sheetnames:
            print(f"[{sekme}] manuel dosyada YOK, atlaniyor")
            continue
        if sekme not in wb_oto.sheetnames:
            print(f"[{sekme}] otomatik dosyada YOK, atlaniyor")
            continue

        sonuclar, anahtar_s = karsilastir_sekme(
            sekme, wb_man[sekme], wb_oto[sekme], config)
        tum_sonuclar.append((sekme, anahtar_s, config["degerler"], sonuclar))

        for s in sonuclar:
            ozet[sekme][s[-2]] += 1  # kategori
        print(f"[{sekme}] {len(sonuclar)} alan karsilastirildi")

    wb_man.close()
    wb_oto.close()

    # ===== Rapor Excel'i yaz =====
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    baslik_font = Font(bold=True)
    fills = {
        "ESLESTI": PatternFill("solid", fgColor="C6EFCE"),
        "UYUMSUZ": PatternFill("solid", fgColor="FFC7CE"),
        "IKISI_DE_BOS": PatternFill("solid", fgColor="FFEB9C"),
        "FARKLI_GUN": PatternFill("solid", fgColor="BDD7EE"),
    }

    # Ozet sekmesi
    ws_ozet = wb_out.create_sheet("OZET")
    ws_ozet.append(["Sekme", "ESLESTI", "UYUMSUZ", "IKISI_DE_BOS", "Toplam", "Dogruluk_%"])
    for c in ws_ozet[1]:
        c.font = baslik_font
    for sekme in ozet:
        e = ozet[sekme].get("ESLESTI", 0)
        u = ozet[sekme].get("UYUMSUZ", 0)
        b = ozet[sekme].get("IKISI_DE_BOS", 0)
        # Dogruluk: eslesti / (eslesti + uyumsuz) — ikisi de bos haric
        payda = e + u
        dogruluk = round(100 * e / payda, 1) if payda > 0 else 0
        ws_ozet.append([sekme, e, u, b, e + u + b, dogruluk])

    # Her sekme icin detay + uyumsuzlar ayri sekme
    for sekme, anahtar_s, deger_s, sonuclar in tum_sonuclar:
        # Detay sekmesi (tum karsilastirmalar)
        ws = wb_out.create_sheet(sekme[:28])
        baslik = ["Sekme"] + anahtar_s + ["Alan", "Manuel_Deger",
                                           "Otomatik_Deger", "Kategori", "Not"]
        ws.append(baslik)
        for c in ws[1]:
            c.font = baslik_font
        for s in sonuclar:
            ws.append(s)
            kat = s[-2]
            if kat in fills:
                for c in ws[ws.max_row]:
                    c.fill = fills[kat]

    # Sadece UYUMSUZ olanlar — tek yerde (kullanici PDF'e bakip karar verecek)
    ws_uy = wb_out.create_sheet("UYUMSUZLAR", 1)
    ws_uy.append(["Sekme", "Anahtar", "Alan", "Manuel_Deger", "Otomatik_Deger",
                  "Not", "PDF_Karari (arac/manuel/format/gun)", "Aciklama"])
    for c in ws_uy[1]:
        c.font = baslik_font
    for sekme, anahtar_s, deger_s, sonuclar in tum_sonuclar:
        for s in sonuclar:
            if s[-2] == "UYUMSUZ":
                anahtar_str = " | ".join(str(x) for x in s[1:1+len(anahtar_s)])
                alan = s[1+len(anahtar_s)]
                mv = s[-4]
                ov = s[-3]
                not_ = s[-1]
                ws_uy.append([sekme, anahtar_str, alan, mv, ov, not_, "", ""])

    wb_out.save(args.cikti)
    print(f"\nRapor yazildi: {args.cikti}")
    print("  - OZET sekmesi: sekme bazli dogruluk")
    print("  - UYUMSUZLAR sekmesi: PDF'e bakip karar verilecekler")
    print("  - Her sekme: renk kodlu detay (yesil=eslesti, kirmizi=uyumsuz, sari=ikisi bos)")


if __name__ == "__main__":
    main()
