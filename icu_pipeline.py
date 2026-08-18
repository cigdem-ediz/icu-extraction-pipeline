"""
YBÜ PDF → Excel Pipeline  v6.0
================================
TAMAMEN LOKAL — internet yok, veri dışarı çıkmaz.

Kullanım:
    python icu_pipeline.py --klasor /hasta/pdfleri --cikti sonuc.xlsx

Dosya adlandırma kuralı (her hasta için):
    100001_E1.pdf   → Epikriz ana metin
    100001_E2.pdf   → Günlük YBÜ formu (GKS, sepsis)
    100001_LAB.pdf  → Laboratuvar özeti
    100001_M1.pdf   → Mikrobiyoloji 1 (M1..M9)

Protokol no dosya adında bulunmalı.
Prefix olabilir: 900001_100001_E1.pdf formatı da desteklenir.

Kurulum (bir kez):
    pip install pdfplumber openpyxl
"""

import os, re, sys, argparse, logging
import csv, time, statistics  # [v6] hasta basina sure logu icin
logging.getLogger('pdfplumber').setLevel(logging.ERROR)
logging.getLogger('pdfminer').setLevel(logging.ERROR)
from pathlib import Path
from datetime import date, datetime
from collections import defaultdict

import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ═══════════════════════════════════════════════════════════════════════════════
# GENEL YARDIMCILAR
# ═══════════════════════════════════════════════════════════════════════════════

def normalize(text):
    if not text: return ""
    return re.sub(r'\s+', ' ', str(text).strip())

def to_float(s):
    s = str(s).strip().rstrip(',').rstrip('.')
    try: return float(s.replace(',', '.'))
    except: return None

def parse_date(s):
    s = str(s).strip()
    m = re.match(r'^(\d{1,2})\.(\d{1,2})\.(\d{4})$', s)
    if m:
        try: return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except: pass
    return None

def get_pdf_text(path):
    pages = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            t = page.extract_text()
            if t: pages.append(t)
    return "\n".join(pages)

# ═══════════════════════════════════════════════════════════════════════════════
# DOSYA GRUPLAMA
# ═══════════════════════════════════════════════════════════════════════════════

def get_hastane_adi(klasor):
    """
    Klasör adından hastane adını çıkar.
    pdfs_Istanbul  → Istanbul
    pdfs_Ankara    → Ankara
    pdfs           → (boş)
    istanbul_pdfs  → istanbul
    Sadece_pdfs    → (boş)
    """
    klasor_adi = Path(klasor).name
    # "pdfs_Istanbul" veya "Istanbul_pdfs" formatı
    m = re.match(r'^pdfs[_\-\s](.+)$', klasor_adi, re.IGNORECASE)
    if m:
        return m.group(1).strip()
    m = re.match(r'^(.+)[_\-\s]pdfs$', klasor_adi, re.IGNORECASE)
    if m:
        return m.group(1).strip()
    # Klasör adı doğrudan hastane adıysa (pdfs değilse)
    if klasor_adi.lower() not in ('pdfs', 'pdf', '.', 'uploads', 'upload'):
        return klasor_adi
    return ""

def group_pdfs(klasor):
    """
    Klasördeki PDF'leri protokol no'ya göre grupla.
    Desteklenen formatlar:
      100001_E1.pdf
      100001 E1.pdf
      900001_100001_E1.pdf
      100001-E1.pdf
    """
    groups = defaultdict(dict)
    unmatched = []
    SEP = r'[_\-\s]'
    for f in sorted(Path(klasor).glob("*.pdf")):
        name = f.stem
        m = re.search(SEP + r'?(\d{5,8})' + SEP + r'(E1|E2|LAB|M\d+)$', name, re.IGNORECASE)
        if not m:
            m = re.search(r'^(\d{5,8})' + SEP + r'(E1|E2|LAB|M\d+)', name, re.IGNORECASE)
        if not m:
            m = re.search(r'(\d{5,8})\s+(E1|E2|LAB|M\d+)', name, re.IGNORECASE)
        if m:
            groups[m.group(1)][m.group(2).upper()] = str(f)
        else:
            unmatched.append(f.name)
    if unmatched:
        print(f"  [UYARI] Tanınamayan dosyalar: {unmatched}")
    return groups

# ═══════════════════════════════════════════════════════════════════════════════
# E1 EPİKRİZ PARSER
# ═══════════════════════════════════════════════════════════════════════════════

def parse_e1(path):
    text = get_pdf_text(path)
    d = {}

    def find(pat, grp=1, flags=re.IGNORECASE):
        m = re.search(pat, text, flags)
        return normalize(m.group(grp)) if m else ""

    # Kimlik
    # Hastane adı: "ÖZEL ... HASTANESİ" gibi (kurum adı çalışma anında PDF'ten çıkarılır)
    m_h = re.search(r'^(.{5,60}HASTANESİ[^\n]*)', text, re.MULTILINE|re.IGNORECASE)
    d['hastane'] = normalize(m_h.group(1))[:60] if m_h else ""
    d['ad_soyad']     = find(r'Adı\s*-\s*Soyadı\s*:\s*(.+?)(?:Prot|$)')
    m = re.search(r'Prot[^\d]*(\d+)\s*/\s*(\d+)', text)
    d['protokol_no']  = m.group(1) if m else ""
    d['dosya_no']     = m.group(2) if m else ""
    d['tc']           = find(r'TC Kimlik No\s*:\s*(\d+)')
    m = re.search(r'Cinsiyeti\s*-\s*Yaş\s*:\s*(ERKEK|KADIN|E|K)\s*-\s*(\d+)', text, re.IGNORECASE)
    d['cinsiyet']     = ("E" if m.group(1).upper() in ("ERKEK","E") else "K") if m else ""
    d['yas']          = int(m.group(2)) if m else None
    m = re.search(r'D\.?Tarihi\s*[:\s]+(\d{1,2}\.\d{1,2}\.\d{4})', text)
    d['dogum_tarihi'] = parse_date(m.group(1)) if m else None
    # Dogum tarihi yoksa bos birak - E2'den alinacak
    # (yas'tan hesaplama hatali olabilir)
    m = re.search(r'Yatış Tarihi\s*:\s*(\d{1,2}\.\d{1,2}\.\d{4})', text)
    d['yatis_tarihi'] = parse_date(m.group(1)) if m else None
    m = re.search(r'Çıkış Tarihi\s*:\s*(\d{1,2}\.\d{1,2}\.\d{4})', text)
    d['cikis_tarihi'] = parse_date(m.group(1)) if m else None
    d['yatis_suresi'] = (d['cikis_tarihi'] - d['yatis_tarihi']).days if (d['yatis_tarihi'] and d['cikis_tarihi']) else None
    d['taburcu_sekli']= find(r'Taburcu Şekli\s*:\s*(.+?)(?:\n|Doktor)')
    d['kurum']        = find(r'Kurum\s*:\s*(SGK|ÖZEL|SUT|GSS|[^\n/]+)')
    d['bolum']        = find(r'Bölüm Adı\s*:\s*(.+?)(?:Yatış|Adresi)')


    # Tanılar — Kesin Tanı satırlarından kod ve isim ayrı çek
    tani_satirlari = re.findall(
        r'(?:Kesin|Ön) Tanı\s*:\s*(.+?)(?=(?:Kesin|Ön) Tanı|Şikayeti)',
        text, re.DOTALL
    )
    icd_kodlar = []
    hastalik_isimleri = []
    for t in tani_satirlari:
        t = normalize(t)
        km = re.match(r'([A-Z]\d+\.?\d*)\s*-\s*(.+)', t)
        if km:
            icd_kodlar.append(km.group(1).strip())
            isim = re.sub(r'\s*\(.*?\)\s*', ' ', km.group(2)).strip()
            isim = re.sub(r'\s+', ' ', isim).strip()
            hastalik_isimleri.append(isim)
        else:
            hastalik_isimleri.append(t)
    d['icd_kodlari'] = " / ".join(icd_kodlar)
    d['tanilar']     = " | ".join(hastalik_isimleri)
    m = re.search(r'Şikayeti\s*:\s*(.+?)(?:Hikayesi|Fiziki)', text, re.DOTALL)
    d['sikayet']      = normalize(m.group(1))[:300] if m else ""

    # Klinik
    m = re.search(r'Apache II skoru[:\s]*(\d+)', text, re.IGNORECASE)
    d['apache_ii_0']      = int(m.group(1)) if m else None
    m = re.search(r'tanısal ölüm oranı[\s\S]*?%(\d+[\.,]?\d*)', text, re.IGNORECASE)
    if not m:
        m = re.search(r'ölüm oranı\s*[:\s%]*(\d+[\.,]?\d*)', text, re.IGNORECASE)
    d['beklenen_olum_0']  = float(m.group(1).replace(',','.')) if m else None
    # Ventilasyon
    if re.search(r'trakeostomi|trakeo', text, re.IGNORECASE):
        d['ventilasyon'] = "Trakeostomi"
    elif re.search(r'entübe|entübasyon', text, re.IGNORECASE):
        d['ventilasyon'] = "Entübe"
    else:
        d['ventilasyon'] = "Yok"

    # SVK
    # [v6] 'kateterizasyon' tek basina yanlis pozitif uretiyordu: uriner/mesane
    # kateterizasyonu da esliyordu. Artik yalnizca santral venoz erisim ifadeleri.
    # Bkz. Kisitliliklar: SVK tespiti anahtar kelime tabanlidir.
    _SVK_RE = r'subklavyen|santral\s*ven|santral\s*kateter|SVK\b|juguler|femoral\s*ven|port\s*kateter'
    d['svk'] = "EVET" if re.search(_SVK_RE, text, re.IGNORECASE) else "HAYIR"

    # Sepsis (demografide tutulabilir — klinik sheet'e E2'den gidiyor)
    d['sepsis']     = "EVET" if re.search(r'pnömosepsis|sepsis', text, re.IGNORECASE) else "HAYIR"
    d['septik_sok'] = "EVET" if re.search(r'septik şok|vazopressör', text, re.IGNORECASE) else "HAYIR"

    # Enfeksiyon kaynağı
    enf_kaynak = []
    if re.search(r'pnömosepsis|pulmoner|pnömoni|akciğer enfek', text, re.IGNORECASE):
        enf_kaynak.append("Pulmoner")
    if re.search(r'üriner sistem enfek|idrar yolu|ütf', text, re.IGNORECASE):
        enf_kaynak.append("Üriner")
    if re.search(r'intraabdominal|abdominal enfek|peritonit', text, re.IGNORECASE):
        enf_kaynak.append("İntraabdominal")
    if re.search(r'katater ile ilişkili|kateter enfek|CLABSI|kan yolu enfek', text, re.IGNORECASE):
        enf_kaynak.append("Katater İlişkili")
    if re.search(r'yara enfek|cerrahi alan', text, re.IGNORECASE):
        enf_kaynak.append("Yara")
    d['enfeksiyon_kaynagi'] = " | ".join(enf_kaynak) if enf_kaynak else "Belirsiz" 
    # Beslenme — başlangıç, son durum, geçiş günü
    # Günlük notları tara
    gun_blocks = re.split(r'(\d{1,2}\.\d{2}\.\d{4})\s+\d+\.GÜN', text)
    beslenme_gunluk = {}  # {gun_no: tip}  tip: "Parenteral" / "Enteral" / "Oral"
    
    yatis = d.get('yatis_tarihi')
    j = 1
    while j < len(gun_blocks) - 1:
        tarih_str = gun_blocks[j]
        icerik    = gun_blocks[j+1]
        tarih = parse_date(tarih_str)
        gun_no = (tarih - yatis).days if (tarih and yatis) else None
        if gun_no is not None:
            if re.search(r'enteral beslenme\s*\(\+\)', icerik, re.IGNORECASE):
                beslenme_gunluk[gun_no] = "Enteral"
            elif re.search(r'oral beslenme|oral alım', icerik, re.IGNORECASE):
                beslenme_gunluk[gun_no] = "Oral"
            elif re.search(r'parenteral beslenme\s*\(\+\)', icerik, re.IGNORECASE):
                beslenme_gunluk[gun_no] = "Parenteral"
            # Enteral(-) varsa parenteral
            elif re.search(r'enteral beslenme\s*\(\-\)', icerik, re.IGNORECASE):
                beslenme_gunluk[gun_no] = "Parenteral"
        j += 2

    if beslenme_gunluk:
        sorted_guns = sorted(beslenme_gunluk.items())
        d['beslenme_baslangic'] = sorted_guns[0][1]
        d['beslenme_son']       = sorted_guns[-1][1]
        # Geçiş günü: başlangıç tipi değiştiği ilk gün
        d['beslenme_gecis_gunu'] = None
        ilk_tip = sorted_guns[0][1]
        for gun_no, tip in sorted_guns[1:]:
            if tip != ilk_tip:
                d['beslenme_gecis_gunu'] = gun_no
                break
    else:
        d['beslenme_baslangic']  = ""
        d['beslenme_son']        = ""
        d['beslenme_gecis_gunu'] = None
    if re.search(r'SRRT|hemodiyaliz|\bHD\b', text, re.IGNORECASE):
        d['renal_durum'] = "SRRT/HD"
    elif re.search(r'akut böbrek yetmezliği|AKI', text, re.IGNORECASE):
        d['renal_durum'] = "AKI"
    else:
        d['renal_durum'] = "Normal"

    return d

# ─── E1: Antimikrobial kurslar ────────────────────────────────────────────


# ── AMR Referans Tablosu ──────────────────────────────────────────────────────
# Bu dosya icu_pipeline.py'ye entegre edilecek AMR modülü

# ══════════════════════════════════════════════════════════════════
# AMR REFERENCE TABLES (WHO ATC/DDD 2026-01-20 + TİTCK, pharmacist-verified)
# Ürün adı → ATC eşleştirmesi bir klinik eczacı tarafından doğrulandı.
# ══════════════════════════════════════════════════════════════════

# Tablo 1: Ürün adı (normalize) → ATC kodu
URUN_ATC = {
    "alfasid": "J01CR01",
    "amikaver": "J01GB06",
    "bactrim": "J01EE01",
    "candisept": "J02AC01",
    "cefperazon": "J01DD62",
    "cilapem": "J01DH51",
    "colimycin": "J01XB01",
    "colitim": "J01XB01",
    "flotic": "J01MA02",
    "fluject": "J02AC01",
    "flukol": "J02AC01",
    "fosit": "J01XX01",
    "genthaver": "J01GB03",
    "iesetum": "J01DD02",
    "klacid": "J01FA09",
    "levoxipolin": "J01MA12",
    "linezosel": "J01XX08",
    "merosid": "J01DH02",
    "moxacin": "J01MA14",
    "mrsacin": "J01AA12",
    "multiflexlinezosel": "J01XX08",
    "nevakson": "J01DD04",
    "polgyl": "J01XD01",
    "rif": "J04AB03",
    "selfleksciprasel": "J01MA02",
    "targocid": "J01XA02",
    "tazeracin": "J01CR05",
    "tazoject": "J01CR05",
    "tygex": "J01AA12",
    "urocare": "J01XX01",
    "vancomax": "J01XA01",

    # ═══════════════════════════════════════════════════════════════════════
    # EK: YBU epikrizinde HEKIMLERIN kullandigi kisa/marka adlari.
    # Kaynak PDF eczane urun adini degil, hekimin yazdigi kisa adi icerir
    # (orn. "Meronem", "Tygacil"); bu nedenle yukaridaki eczane adlari yetmiyordu.
    # Bu eslesmeler daha once olu _ILAC_KISALTMA / _ILAC_ATC_HARITA tablolarinda
    # duruyordu, ancak hicbiri lookup_amr()e bagli degildi -> ATC/AWaRe/DDD bos kaliyordu.
    # Tum eslesmeler bir klinik eczaci tarafindan dogrulanmistir.
    # ═══════════════════════════════════════════════════════════════════════
    "meronem":     "J01DH02",   # Meropenem
    "tygacil":     "J01AA12",   # Tigecycline
    "tavanic":     "J01MA12",   # Levofloxacin
    "cipro":       "J01MA02",   # Ciprofloxacin
    "avelox":      "J01MA14",   # Moxifloxacin
    "fortum":      "J01DD02",   # Ceftazidime
    "lumen":       "J02AX04",   # Caspofungin
    "vanco":       "J01XA01",   # Vancomycin
    "amikasin":    "J01GB06",   # Amikacin
    "kolimycin":   "J01XB01",   # Colistin
    "kolimisin":   "J01XB01",   # Colistin
    "novosef":     "J01DD04",   # Ceftriaxone
    "rocephin":    "J01DD04",   # Ceftriaxone
    "sulparazon":  "J01DD62",   # Cefoperazone/sulbactam
    "nodizil":     "J01XX08",   # Linezolid
    "tazocin":     "J01CR05",   # Piperacillin/tazobactam
    "tienam":      "J01DH51",   # Imipenem/cilastatin
    "cilanem":     "J01DH51",   # Imipenem/cilastatin

    # ── EK-2: Tam kohort (206 hasta) calistirmasinda cikan YAZIM VARYANTLARI ──
    # Kaynak belgelerde ayni etken madde farkli/hatali yazimlarla gecebiliyor.
    # Tumu klinik eczaci tarafindan dogrulanmistir.
    "colimiycin":  "J01XB01",   # Colimycin yazim varyanti -> Colistin
    "sulparezon":  "J01DD62",   # Sulperazon -> Cefoperazone/sulbactam
    "tianem":      "J01DH51",   # Tienam -> Imipenem/cilastatin
    "vankomycin":  "J01XA01",   # Vancomycin
    "vankomisin":  "J01XA01",   # Vancomycin
    "gentamycin":  "J01GB03",   # Gentamicin
    "genta":       "J01GB03",   # Gentamicin (kisaltma)
    "gentamisin":  "J01GB03",   # Gentamicin
    "flagly":      "J01XD01",   # Flagyl -> Metronidazole
    "flagyl":      "J01XD01",   # Metronidazole
    "zidim":       "J01DD02",   # Ceftazidime
    "seffur":      "J01DC02",   # Cefuroxime
    "sulcid":      "J01CR01",   # Ampicillin/sulbactam (IV)
    "duocid":      "J01CR04",   # Sultamicillin — ORAL tablet (IV->oral step-down)
    "doucid":      "J01CR04",   # Duocid yazim varyanti (kaynakta "Doucid")
    "sultamisilin":"J01CR04",   # Sultamicillin
}

# Tablo 2: ATC → (etken_madde, AWaRe, DDD_g, uygulama_yolu)
ATC_REFERENCE = {
    "J01AA01": ("Demeclocycline", "Watch", None, ""),
    "J01AA02": ("Doxycycline", "Access", None, ""),
    "J01AA03": ("Chlortetracycline", "Watch", None, ""),
    "J01AA04": ("Lymecycline", "Watch", None, ""),
    "J01AA05": ("Metacycline", "Watch", None, ""),
    "J01AA06": ("Oxytetracycline", "Watch", None, ""),
    "J01AA07": ("Tetracycline", "Access", None, ""),
    "J01AA08": ("Minocycline", "Reserve", None, ""),
    "J01AA09": ("Rolitetracycline", "Watch", None, ""),
    "J01AA10": ("Penimepicycline", "Watch", None, ""),
    "J01AA11": ("Clomocycline", "Watch", None, ""),
    "J01AA12": ("Tigecycline", "Reserve", 0.1, "P"),
    "J01AA13": ("Eravacycline", "Reserve", None, ""),
    "J01AA14": ("Sarecycline", "Watch", None, ""),
    "J01AA15": ("Omadacycline", "Reserve", None, ""),
    "J01BA01": ("Chloramphenicol", "Access", None, ""),
    "J01BA02": ("Thiamphenicol", "Access", None, ""),
    "J01CA01": ("Ampicillin", "Access", None, ""),
    "J01CA02": ("Pivampicillin", "Access", None, ""),
    "J01CA03": ("Carbenicillin", "Watch", None, ""),
    "J01CA04": ("Amoxicillin", "Access", None, ""),
    "J01CA05": ("Carindacillin", "Watch", None, ""),
    "J01CA06": ("Bacampicillin", "Access", None, ""),
    "J01CA07": ("Epicillin", "Access", None, ""),
    "J01CA08": ("Pivmecillinam", "Access", None, ""),
    "J01CA09": ("Azlocillin", "Watch", None, ""),
    "J01CA10": ("Mezlocillin", "Watch", None, ""),
    "J01CA11": ("Mecillinam", "Access", None, ""),
    "J01CA12": ("Piperacillin", "Watch", None, ""),
    "J01CA13": ("Ticarcillin", "Watch", None, ""),
    "J01CA14": ("Metampicillin", "Access", None, ""),
    "J01CA15": ("Talampicillin", "Access", None, ""),
    "J01CA16": ("Sulbenicillin", "Watch", None, ""),
    "J01CA17": ("Temocillin", "Watch", None, ""),
    "J01CA18": ("Hetacillin", "Access", None, ""),
    "J01CA19": ("Aspoxicillin", "Watch", None, ""),
    "J01CE01": ("Benzylpenicillin", "Access", None, ""),
    "J01CE02": ("Phenoxymethylpenicillin", "Access", None, ""),
    "J01CE03": ("Propicillin", "Access", None, ""),
    "J01CE04": ("Azidocillin", "Access", None, ""),
    "J01CE05": ("Pheneticillin", "Watch", None, ""),
    "J01CE06": ("Penamecillin", "Access", None, ""),
    "J01CE07": ("Clometocillin", "Access", None, ""),
    "J01CE08": ("Benzathine-benzylpenicillin", "Access", None, ""),
    "J01CE09": ("Procaine-benzylpenicillin", "Access", None, ""),
    "J01CF01": ("Dicloxacillin", "Access", None, ""),
    "J01CF02": ("Cloxacillin", "Access", None, ""),
    "J01CF03": ("Meticillin", "Access", None, ""),
    "J01CF04": ("Oxacillin", "Access", None, ""),
    "J01CF05": ("Flucloxacillin", "Access", None, ""),
    "J01CF06": ("Nafcillin", "Access", None, ""),
    "J01CG01": ("Sulbactam", "Access", None, ""),
    "J01CG02": ("Tazobactam", "Watch", None, ""),
    "J01CR01": ("Ampicillin/sulbactam", "Access", 6, "P"),
    "J01CR02": ("Amoxicillin/clavulanic-acid", "Access", None, ""),
    "J01CR04": ("Sultamicillin", "Access", 1.5, "O"),
    "J01CR05": ("Piperacillin/tazobactam", "Watch", 14, "P"),
    "J01DB01": ("Cefalexin", "Access", None, ""),
    "J01DB02": ("Cefaloridine", "Access", None, ""),
    "J01DB03": ("Cefalotin", "Access", None, ""),
    "J01DB04": ("Cefazolin", "Access", None, ""),
    "J01DB05": ("Cefadroxil", "Access", None, ""),
    "J01DB06": ("Cefazedone", "Access", None, ""),
    "J01DB07": ("Cefatrizine", "Access", None, ""),
    "J01DB08": ("Cefapirin", "Access", None, ""),
    "J01DB09": ("Cefradine", "Access", None, ""),
    "J01DB10": ("Cefacetrile", "Access", None, ""),
    "J01DB11": ("Cefroxadine", "Access", None, ""),
    "J01DB12": ("Ceftezole", "Access", None, ""),
    "J01DC01": ("Cefoxitin", "Watch", None, ""),
    "J01DC02": ("Cefuroxime", "Watch", 3.0, "P"),
    "J01DC03": ("Cefamandole", "Watch", None, ""),
    "J01DC04": ("Cefaclor", "Watch", None, ""),
    "J01DC05": ("Cefotetan", "Watch", None, ""),
    "J01DC06": ("Cefonicid", "Watch", None, ""),
    "J01DC07": ("Cefotiam", "Watch", None, ""),
    "J01DC08": ("Loracarbef", "Watch", None, ""),
    "J01DC09": ("Cefmetazole", "Watch", None, ""),
    "J01DC10": ("Cefprozil", "Watch", None, ""),
    "J01DC11": ("Ceforanide", "Watch", None, ""),
    "J01DC12": ("Cefminox", "Watch", None, ""),
    "J01DC13": ("Cefbuperazone", "Watch", None, ""),
    "J01DC14": ("Flomoxef", "Watch", None, ""),
    "J01DD01": ("Cefotaxime", "Watch", None, ""),
    "J01DD02": ("Ceftazidime", "Watch", 4, "P"),
    "J01DD03": ("Cefsulodin", "Watch", None, ""),
    "J01DD04": ("Ceftriaxone", "Watch", 2, "P"),
    "J01DD05": ("Cefmenoxime", "Watch", None, ""),
    "J01DD06": ("Latamoxef", "Watch", None, ""),
    "J01DD07": ("Ceftizoxime", "Watch", None, ""),
    "J01DD08": ("Cefixime", "Watch", None, ""),
    "J01DD09": ("Cefodizime", "Watch", None, ""),
    "J01DD10": ("Cefetamet-pivoxil", "Watch", None, ""),
    "J01DD11": ("Cefpiramide", "Watch", None, ""),
    "J01DD12": ("Cefoperazone", "Watch", None, ""),
    "J01DD13": ("Cefpodoxime-proxetil", "Watch", None, ""),
    "J01DD14": ("Ceftibuten", "Watch", None, ""),
    "J01DD15": ("Cefdinir", "Watch", None, ""),
    "J01DD16": ("Cefditoren-pivoxil", "Watch", None, ""),
    "J01DD17": ("Cefcapene-pivoxil", "Watch", None, ""),
    "J01DD18": ("Cefteram-pivoxil", "Watch", None, ""),
    "J01DD52": ("Ceftazidime/avibactam", "Reserve", None, ""),
    "J01DD62": ("Cefoperazone/sulbactam", "N/A (not in WHO AWaRe list)", 4, "P"),
    "J01DE01": ("Cefepime", "Watch", None, ""),
    "J01DE02": ("Cefpirome", "Watch", None, ""),
    "J01DE03": ("Cefozopran", "Watch", None, ""),
    "J01DF01": ("Aztreonam", "Reserve", None, ""),
    "J01DF02": ("Carumonam", "Reserve", None, ""),
    "J01DH02": ("Meropenem", "Watch", 3, "P"),
    "J01DH03": ("Ertapenem", "Watch", None, ""),
    "J01DH04": ("Doripenem", "Watch", None, ""),
    "J01DH05": ("Biapenem", "Watch", None, ""),
    "J01DH06": ("Tebipenem", "Watch", None, ""),
    "J01DH51": ("Imipenem/cilastatin", "Watch", 2, "P"),
    "J01DH52": ("Meropenem/vaborbactam", "Reserve", None, ""),
    "J01DH55": ("Panipenem", "Watch", None, ""),
    "J01DH56": ("Imipenem/cilastatin/relebactam", "Reserve", None, ""),
    "J01DI01": ("Ceftobiprole-medocaril", "Reserve", None, ""),
    "J01DI02": ("Ceftaroline-fosamil", "Reserve", None, ""),
    "J01DI03": ("Faropenem", "Reserve", None, ""),
    "J01DI04": ("Cefiderocol", "Reserve", None, ""),
    "J01DI54": ("Ceftolozane/tazobactam", "Reserve", None, ""),
    "J01EA01": ("Trimethoprim", "Access", None, ""),
    "J01EA02": ("Brodimoprim", "Access", None, ""),
    "J01EA03": ("Iclaprim", "Reserve", None, ""),
    "J01EB01": ("Sulfaisodimidine", "Access", None, ""),
    "J01EB02": ("Sulfamethizole", "Access", None, ""),
    "J01EB03": ("Sulfadimidine", "Access", None, ""),
    "J01EB04": ("Sulfapyridine", "Access", None, ""),
    "J01EB05": ("Sulfafurazole", "Access", None, ""),
    "J01EB06": ("Sulfanilamide", "Access", None, ""),
    "J01EB07": ("Sulfathiazole", "Access", None, ""),
    "J01EB08": ("Sulfathiourea", "Access", None, ""),
    "J01EC01": ("Sulfamethoxazole", "Access", None, ""),
    "J01EC02": ("Sulfadiazine", "Access", None, ""),
    "J01EC03": ("Sulfamoxole", "Access", None, ""),
    "J01ED01": ("Sulfadimethoxine", "Access", None, ""),
    "J01ED02": ("Sulfalene", "Access", None, ""),
    "J01ED03": ("Sulfametomidine", "Access", None, ""),
    "J01ED04": ("Sulfametoxydiazine", "Access", None, ""),
    "J01ED05": ("Sulfamethoxypyridazine", "Access", None, ""),
    "J01ED06": ("Sulfaperin", "Access", None, ""),
    "J01ED07": ("Sulfamerazine", "Access", None, ""),
    "J01ED08": ("Sulfaphenazole", "Access", None, ""),
    "J01ED09": ("Sulfamazone", "Access", None, ""),
    "J01EE01": ("Sulfamethoxazole/trimethoprim", "Access", 1.92, "P"),
    "J01EE02": ("Sulfadiazine/trimethoprim", "Access", None, ""),
    "J01EE03": ("Sulfametrole/trimethoprim", "Access", None, ""),
    "J01EE04": ("Sulfamoxole/trimethoprim", "Access", None, ""),
    "J01EE05": ("Sulfadimidine/trimethoprim", "Access", None, ""),
    "J01EE06": ("Sulfadiazine/tetroxoprim", "Access", None, ""),
    "J01EE07": ("Sulfamerazine/trimethoprim", "Access", None, ""),
    "J01FA01": ("Erythromycin", "Watch", None, ""),
    "J01FA02": ("Spiramycin", "Watch", None, ""),
    "J01FA03": ("Midecamycin", "Watch", None, ""),
    "J01FA05": ("Oleandomycin", "Watch", None, ""),
    "J01FA06": ("Roxithromycin", "Watch", None, ""),
    "J01FA07": ("Josamycin", "Watch", None, ""),
    "J01FA08": ("Troleandomycin", "Watch", None, ""),
    "J01FA09": ("Clarithromycin", "Watch", 1, "P"),
    "J01FA10": ("Azithromycin", "Watch", None, ""),
    "J01FA11": ("Miocamycin", "Watch", None, ""),
    "J01FA12": ("Rokitamycin", "Watch", None, ""),
    "J01FA13": ("Dirithromycin", "Watch", None, ""),
    "J01FA14": ("Flurithromycin", "Watch", None, ""),
    "J01FA15": ("Telithromycin", "Watch", None, ""),
    "J01FA16": ("Solithromycin", "Watch", None, ""),
    "J01FF01": ("Clindamycin", "Access", None, ""),
    "J01FF02": ("Lincomycin", "Watch", None, ""),
    "J01FG01": ("Pristinamycin", "Watch", None, ""),
    "J01FG02": ("Dalfopristin/quinupristin", "Reserve", None, ""),
    "J01GA01": ("Streptomycin", "Watch", None, ""),
    "J01GA02": ("Streptoduocin", "Watch", None, ""),
    "J01GB01": ("Tobramycin", "Watch", None, ""),
    "J01GB03": ("Gentamicin", "Access", 0.24, "P"),
    "J01GB04": ("Kanamycin", "Watch", None, ""),
    "J01GB05": ("Neomycin", "Watch", None, ""),
    "J01GB06": ("Amikacin", "Access", 1, "P"),
    "J01GB07": ("Netilmicin", "Watch", None, ""),
    "J01GB08": ("Sisomicin", "Watch", None, ""),
    "J01GB09": ("Dibekacin", "Watch", None, ""),
    "J01GB10": ("Ribostamycin", "Watch", None, ""),
    "J01GB11": ("Isepamicin", "Watch", None, ""),
    "J01GB12": ("Arbekacin", "Watch", None, ""),
    "J01GB13": ("Bekanamycin", "Watch", None, ""),
    "J01GB14": ("Plazomicin", "Reserve", None, ""),
    "J01MA01": ("Ofloxacin", "Watch", None, ""),
    "J01MA02": ("Ciprofloxacin", "Watch", 0.8, "P"),
    "J01MA03": ("Pefloxacin", "Watch", None, ""),
    "J01MA04": ("Enoxacin", "Watch", None, ""),
    "J01MA05": ("Temafloxacin", "Watch", None, ""),
    "J01MA06": ("Norfloxacin", "Watch", None, ""),
    "J01MA07": ("Lomefloxacin", "Watch", None, ""),
    "J01MA08": ("Fleroxacin", "Watch", None, ""),
    "J01MA09": ("Sparfloxacin", "Watch", None, ""),
    "J01MA10": ("Rufloxacin", "Watch", None, ""),
    "J01MA11": ("Grepafloxacin", "Watch", None, ""),
    "J01MA12": ("Levofloxacin", "Watch", 0.5, "P"),
    "J01MA13": ("Trovafloxacin", "Watch", None, ""),
    "J01MA14": ("Moxifloxacin", "Watch", 0.4, "P"),
    "J01MA15": ("Gemifloxacin", "Watch", None, ""),
    "J01MA16": ("Gatifloxacin", "Watch", None, ""),
    "J01MA17": ("Prulifloxacin", "Watch", None, ""),
    "J01MA18": ("Pazufloxacin", "Watch", None, ""),
    "J01MA19": ("Garenoxacin", "Watch", None, ""),
    "J01MA21": ("Sitafloxacin", "Watch", None, ""),
    "J01MA22": ("Tosufloxacin", "Watch", None, ""),
    "J01MA23": ("Delafloxacin", "Watch", None, ""),
    "J01MA24": ("Levonadifloxacin", "Watch", None, ""),
    "J01MA25": ("Lascufloxacin", "Watch", None, ""),
    "J01MB01": ("Rosoxacin", "Watch", None, ""),
    "J01MB03": ("Piromidic-acid", "Watch", None, ""),
    "J01MB04": ("Pipemidic-acid", "Watch", None, ""),
    "J01MB05": ("Oxolinic-acid", "Watch", None, ""),
    "J01MB06": ("Cinoxacin", "Watch", None, ""),
    "J01MB07": ("Flumequine", "Watch", None, ""),
    "J01MB08": ("Nemonoxacin", "Watch", None, ""),
    "J01XA01": ("Vancomycin", "Watch", 2, "P"),
    "J01XA02": ("Teicoplanin", "Watch", 0.4, "P"),
    "J01XA03": ("Telavancin", "Reserve", None, ""),
    "J01XA04": ("Dalbavancin", "Reserve", None, ""),
    "J01XA05": ("Oritavancin", "Reserve", None, ""),
    "J01XB01": ("Colistin", "Reserve", 9, "P"),
    "J01XB02": ("Polymyxin-B", "Reserve", None, ""),
    "J01XC01": ("Fusidic-acid", "Watch", None, ""),
    "J01XD01": ("Metronidazole", "Access", 1.5, "P"),
    "J01XD02": ("Tinidazole", "Access", None, ""),
    "J01XD03": ("Ornidazole", "Access", None, ""),
    "J01XE01": ("Nitrofurantoin", "Access", None, ""),
    "J01XE02": ("Nifurtoinol", "Access", None, ""),
    "J01XE03": ("Furazidin", "Access", None, ""),
    "J01XX01": ("Fosfomycin", "Reserve", 8, "P"),
    "J01XX03": ("Clofoctol", "Watch", None, ""),
    "J01XX04": ("Spectinomycin", "Access", None, ""),
    "J01XX08": ("Linezolid", "Reserve", 1.2, "P"),
    "J01XX09": ("Daptomycin", "Reserve", None, ""),
    "J01XX11": ("Tedizolid", "Reserve", None, ""),
    "J01XX12": ("Lefamulin", "Reserve", None, ""),
    "J02AC01": ("Fluconazole", "N/A (antifungal)", 0.2, "P"),
    "J02AX04": ("Caspofungin", "N/A (antifungal)", 0.05, "P"),
    "J04AB02": ("Rifampicin", "Watch", None, ""),
    "J04AB03": ("Rifamycin", "Watch", 0.6, "P"),
    "J04AB04": ("Rifabutin", "Watch", None, ""),
}

def normalize_ilac_key(raw):
    import re
    s = raw.lower().strip()
    s = re.sub(r'\(.*?\)', '', s)
    s = re.sub(r'\d+x\d+\s*(?:mg|gr|ml|iu|mcg)?', '', s, flags=re.IGNORECASE)
    s = re.sub(r'\d+\s*(?:mg|gr|ml|iu|mcg)\b', '', s, flags=re.IGNORECASE)
    s = re.sub(r'\s+', ' ', s).strip()
    return s

def _marka_key(ham):
    """Ham ürün adından marka anahtarı çıkar (URUN_ATC ile uyumlu, boşluksuz, Türkçe-normalize)."""
    import re as _re
    s = str(ham).strip().upper()
    # Recete onekini at: "EHU onayi ile Meronem" -> "MERONEM"
    s = _re.sub(r"^\s*(EHU|E\.H\.U\.?)\s*ONAY\w*\s*[İI]?LE\s+", "", s)
    s = _re.sub(r"^\s*(EHU|E\.H\.U\.?)\s+", "", s)
    s = s.replace("İ", "I").replace("Ş", "S").replace("Ğ", "G") \
         .replace("Ü", "U").replace("Ö", "O").replace("Ç", "C").replace("I", "I")
    # ilk dozaj/form kelimesinden öncesini al
    m = _re.search(r"\b(IV|IM|ENJ|FLAKON|FLK|AMPUL|AMP|INF|INFUZYON|SOLUSYON|"
                   r"COZELTI|COZ|LIY|TOZ|SASE|TORBA|MG|ML|GR|G|IU|MCG|LIYOFILIZE|"
                   r"ENJEKTABL|KAPSUL|TABLET|%|/)\b", s)
    if m:
        s = s[:m.start()]
    s = _re.sub(r"\d+", " ", s)
    s = _re.sub(r"[^\w\s]", " ", s)
    s = _re.sub(r"\s+", " ", s).strip().lower().replace(" ", "")
    return s

def lookup_amr(ilac_adi):
    """
    İki aşamalı arama: ürün adı → ATC (URUN_ATC) → (etken, aware, ddd, yol) (ATC_REFERENCE).
    Doğrulanmış sabit tablolar kullanılır; eşleşme yoksa (None,...) döner.
    Geriye dönük uyumluluk için 6'lı tuple döndürür: (etken, atc, aware, ddd, yol, tdm)
    """
    mk = _marka_key(ilac_adi)
    atc = URUN_ATC.get(mk)
    if atc is None and mk:
        # tam kelime bazlı kısmi eşleşme (anahtar, markanın başında/tamamında)
        for uk, uatc in URUN_ATC.items():
            if uk and len(uk) >= 4 and (mk.startswith(uk) or uk.startswith(mk)):
                atc = uatc
                break
    if atc and atc in ATC_REFERENCE:
        etken, aware, ddd, yol = ATC_REFERENCE[atc]
        tdm = atc in ("J01XA01", "J01XA02", "J01XB01", "J01GB06", "J01GB03", "J01XX08")
        return (etken, atc, aware, ddd, yol or "IV", tdm)
    return (None, None, None, None, None, None)
# [KALDIRILDI] parse_vazopressor: cift tanim, sonraki tanim tarafindan eziliyor
# [KALDIRILDI] _enrich_amr_from_pdf: hicbir yerden cagrilmiyor


def parse_antimikrobial(e1_path):
    """E1'den antimikrobiyal kursları çıkar — doz, sıklık, DDD, TDM dahil."""
    from datetime import timedelta
    text = get_pdf_text(e1_path)
    courses = {}

    gun_blocks = re.split(r'(\d{1,2}\.\d{2}\.\d{4})\s+\d+\.GÜN', text)
    i = 1
    while i < len(gun_blocks) - 1:
        tarih = parse_date(gun_blocks[i])
        icerik = re.sub(r'\s+', ' ', gun_blocks[i+1])
        if not tarih:
            i += 2; continue

        for m in re.finditer(
            r'(?:EHU onayıyla|Antifungal tedavi|Antibiyoterapi[^.]*?)\s+'
            r'([A-Za-zÇçĞğİıÖöŞşÜü][A-Za-zÇçĞğİıÖöŞşÜü\s]*)'
            r'\s*(\d+x\d+\s*(?:mg|gr|ml|IU|mcg)?|\d+\s*(?:mg|gr|ml|IU|mcg))?'
            r'\s*\((\d+)\.\s*(?:gün|Gün|GÜN)\)',
            icerik, re.IGNORECASE
        ):
            ilac_raw = m.group(1).strip()
            # Prefix temizle: "EHU onayıyla Avelox" → "Avelox"
            for pfx in ['EHU onayıyla ', 'Antifungal tedavi ', 'ile ', 've ']:
                if ilac_raw.startswith(pfx):
                    ilac_raw = ilac_raw[len(pfx):].strip()
            doz_raw  = (m.group(2) or "").strip()
            gun_no   = int(m.group(3))
            ilac_key = normalize_ilac_key(ilac_raw)
            if not ilac_key or len(ilac_key) < 2:
                continue
            courses.setdefault(ilac_key, []).append((gun_no, tarih, ilac_raw, doz_raw))
        i += 2

    results = []
    for ilac_key, entries in courses.items():
        entries.sort(key=lambda x: x[0])
        kurs, kurs_no, prev_gun = [], 0, None
        for entry in entries:
            gun_no = entry[0]
            if prev_gun is None or gun_no - prev_gun > 2:
                if kurs:
                    r = _build_amr_kurs(kurs, kurs_no, ilac_key)
                    if r: results.append(r)
                kurs, kurs_no = [], kurs_no + 1
            kurs.append(entry)
            prev_gun = gun_no
        if kurs:
            r = _build_amr_kurs(kurs, kurs_no, ilac_key)
            if r: results.append(r)
    return results

def _build_amr_kurs(entries, kurs_no, ilac_key):
    ilac_raw = entries[0][2]
    doz_raws = [e[3] for e in entries if e[3]]
    doz_str  = doz_raws[0] if doz_raws else ""
    sure     = entries[-1][0] - entries[0][0] + 1

    siklik = doz = ""
    if doz_str:
        m = re.match(r'(\d+)x(\d+\s*(?:mg|gr|ml|IU|mcg)?)', doz_str, re.IGNORECASE)
        if m:
            siklik = f"{m.group(1)}x1"
            doz    = m.group(2).strip()
        else:
            m2 = re.match(r'(\d+\s*(?:mg|gr|ml|IU|mcg))', doz_str, re.IGNORECASE)
            if m2:
                doz = m2.group(1).strip(); siklik = "1x1"

    ref = lookup_amr(ilac_key)
    etken, atc, aware, ddd_ref, ref_yol, tdm = ref

    ddd_hesap = None; ddd_notu = ""
    if ddd_ref and doz:
        try:
            dm = re.search(r'([\d\.]+)\s*(mg|gr|g|ml|IU|mcg)?', doz, re.IGNORECASE)
            if dm:
                val  = float(dm.group(1))
                unit = (dm.group(2) or "mg").lower()
                if unit == "mg": val /= 1000
                elif unit == "mcg": val /= 1_000_000
                elif unit == "iu": val = None
                if val:
                    frek = int(siklik.split("x")[0]) if siklik and "x" in siklik else 1
                    ddd_hesap = round(sure * val * frek / ddd_ref, 2)
        except:
            ddd_notu = "Hesaplanamadı"
    if ddd_hesap is None and not ddd_notu:
        ddd_notu = "Doz bilgisi eksik" if not doz else "Hesaplanamadı"

    if not ilac_raw.strip(): return None
    return {
        "ilac_adi": ilac_raw, "etken_madde": etken or "?",
        "atc": atc or "?", "aware": aware or "?",
        "doz": doz or "?", "siklik": siklik or "?",
        "uygulama_yolu": ref_yol or "IV",
        "baslangic": entries[0][1], "bitis": entries[-1][1],
        "sure_gun": sure, "baslangic_gun_no": entries[0][0] - 1,
        "kurs_no": kurs_no, "ehu_onayi": "EVET",
        "ddd": ddd_hesap, "ddd_notu": ddd_notu,
        "tdm": "EVET" if tdm else "HAYIR",
        "tedavi_turu": "", "uygunluk": "",
        "uygunsuzluk_nedeni": "", "deskalasyon": "", "eskalasyon_nedeni": "",
    }

def parse_vazopressor(e1_path, e2_path=None):
    """
    E1 (ve opsiyonel E2) epikrizden vazopressör kullanım kurslarını çıkar.
    
    Kural:
    - Her gün (+) / (-) / yok → günlük durum listesi oluştur
    - Ardışık (+) günleri = 1 kurs
    - Arada kesinti varsa → yeni kurs
    
    Döner: [{'ilac', 'baslangic', 'bitis', 'sure_gun', 'kurs_no'}, ...]
    """
    from datetime import timedelta

    ILACLAR = {
        'noradrenalin': 'Noradrenalin',
        'adrenalin':    'Adrenalin', 
        'dobutamin':    'Dobutamin',
        'dopamin':      'Dopamin',
    }

    def parse_pdf_vazo(path):
        """PDF'ten günlük vazopressör durumunu çek."""
        with pdfplumber.open(path) as pdf:
            pages = [p.extract_text() or "" for p in pdf.pages]
        text = "\n".join(pages)
        
        gun_blocks = re.split(r'(\d{1,2}\.\d{2}\.\d{4})\s+\d+\.GÜN', text)
        daily = {}  # {date: {ilac: bool}}
        
        i = 1
        while i < len(gun_blocks) - 1:
            tarih = parse_date(gun_blocks[i])
            if not tarih:
                i += 2; continue
            icerik = re.sub(r'\s+', ' ', gun_blocks[i+1])
            
            daily[tarih] = {}
            for key, label in ILACLAR.items():
                # "adrenalin" için: "noradrenalin" içinde geçmemeli
                pat = rf'(?<![a-zçğışöü]){key}\s+infüzyonu'
                m_pos = re.search(pat + r'\s*\(\+\)', icerik, re.IGNORECASE)
                m_neg = re.search(pat + r'\s*\(-\)', icerik, re.IGNORECASE)
                m_any = re.search(pat, icerik, re.IGNORECASE)
                
                if m_pos:
                    daily[tarih][key] = True
                elif m_neg:
                    daily[tarih][key] = False
                elif m_any:
                    daily[tarih][key] = True  # (+) yazılmamış ama infüzyon var
            i += 2
        return daily

    # E1'den al
    daily = parse_pdf_vazo(e1_path)
    
    # E2 varsa eksik günleri tamamla
    if e2_path:
        try:
            daily2 = parse_pdf_vazo(e2_path)
            for dt, d2 in daily2.items():
                if dt not in daily:
                    daily[dt] = d2
                else:
                    for ilac, val in d2.items():
                        if ilac not in daily[dt]:
                            daily[dt][ilac] = val
        except:
            pass

    # Kurs tespiti: her ilaç için ardışık günleri grupla
    courses = []
    sorted_dates = sorted(daily.keys())
    
    for key, label in ILACLAR.items():
        # Bu ilaç için günlük durum: True / False / None
        status = {}
        for dt in sorted_dates:
            status[dt] = daily[dt].get(key, None)
        
        # Ardışık True günleri → kurs
        kurs_no = 0
        in_course = False
        start = None
        prev_dt = None
        
        for dt in sorted_dates:
            active = status[dt]
            
            if active is True:
                if not in_course:
                    # Yeni kurs başla
                    # Önceki günde False veya None → kesinlikle yeni kurs
                    in_course = True
                    start = dt
                    kurs_no += 1
                # Devam
                prev_dt = dt
            
            elif active is False:
                if in_course:
                    # Kurs bitti
                    courses.append({
                        'ilac':       label,
                        'baslangic':  start,
                        'bitis':      prev_dt,
                        'sure_gun':   (prev_dt - start).days + 1,
                        'kurs_no':    kurs_no,
                    })
                    in_course = False
                    start = None
                prev_dt = dt
            
            else:
                # None: o gün not yok → kesinti sayılır
                if in_course:
                    # Günlük kayıt yoksa kursu kapat
                    courses.append({
                        'ilac':       label,
                        'baslangic':  start,
                        'bitis':      prev_dt,
                        'sure_gun':   (prev_dt - start).days + 1,
                        'kurs_no':    kurs_no,
                    })
                    in_course = False
                    start = None
                prev_dt = dt
        
        # Dosya bitiminde açık kurs
        if in_course and prev_dt:
            courses.append({
                'ilac':       label,
                'baslangic':  start,
                'bitis':      prev_dt,
                'sure_gun':   (prev_dt - start).days + 1,
                'kurs_no':    kurs_no,
            })

    return courses

def parse_e2(path):
    text = get_pdf_text(path)
    records = []
    # E2'den dogum tarihi cek (bazi formatlarda buradadir)
    dogum_e2 = None
    m_d = re.search(r'D\.?Tarihi\s*[:\s]+(\d{1,2}\.\d{1,2}\.\d{4})', text)
    if not m_d:
        m_d = re.search(r'Do.um Tarihi.*?:\s*(\d{1,2}\.\d{2}\.\d{4})', text)
    if m_d: dogum_e2 = parse_date(m_d.group(1))

    blocks = re.split(r'İzlem Tarihi\s*:', text)
    for block in blocks[1:]:
        tarih_m = re.match(r'\s*(\d{1,2}\.\d{1,2}\.\d{4})', block)
        if not tarih_m: continue
        tarih = parse_date(tarih_m.group(1))
        if not tarih: continue
        rec = {'tarih': tarih}
        def ei(pat): 
            m = re.search(pat, block)
            return int(m.group(1)) if m else None
        # GKS: sadece toplam
        goz   = ei(r'Gözler\s*:\s*(\d)')
        sozel = ei(r'Sözel\s*:\s*(\d)')
        motor = ei(r'Motor\s*:\s*(\d)')
        rec['gks_toplam'] = (goz+sozel+motor) if (goz and sozel and motor) else None
        rec['sepsis']     = "EVET" if re.search(r'Sepsis Durumu\s*:\s*EVET', block) else "HAYIR"
        rec['septik_sok'] = "EVET" if re.search(r'Septik Şok\s*:\s*EVET', block)   else "HAYIR"
        m = re.search(r'(\d)\.\s*BASAMAK', block)
        rec['ybu_seviye'] = int(m.group(1)) if m else None
        m = re.search(r'Toplam Apache Skoru\s*:\s*(\d+)', block)
        rec['apache_ii']  = int(m.group(1)) if m else None
        # Beklenen olum - farkli formatlar
        m = re.search(r'Beklenen Ölüm Oranı\s*:\s*(\d+[,\.]?\d*)', block)
        if not m:
            m = re.search(r'Bek\.\s*Ölüm Oranı\s*:\s*(\d+[,\.]?\d*)', block)
        rec['beklenen_olum'] = float(m.group(1).replace(',','.')) if m else None
        # SOFA — varsa al (bazı sistemler yazar)
        m = re.search(r'SOFA\s*(?:skoru|puan)?\s*[:\s]*(\d+)', block, re.IGNORECASE)
        rec['sofa'] = int(m.group(1)) if m else None
        records.append(rec)
    return records, dogum_e2
# [KALDIRILDI] lab_extract_series: cift tanim, sonraki tanim tarafindan eziliyor
# [KALDIRILDI] lab_extract_hemogram: cift tanim, sonraki tanim tarafindan eziliyor
# [KALDIRILDI] e1_extract_kan_gazi: cift tanim, sonraki tanim tarafindan eziliyor
# [KALDIRILDI] e2_extract_lab: cift tanim, sonraki tanim tarafindan eziliyor



# ═══════════════════════════════════════════════════════════════════════════════
# LAB PDF: EVRENSEl PARSER (taşma sayfaları dahil)
# ═══════════════════════════════════════════════════════════════════════════════

def parse_lab_universal(pages):
    """
    LAB PDF'inin tüm sayfalarını işle.
    Taşma sayfaları (başlıksız) otomatik tespit edilir ve
    önceki sayfanın parametreleri ile eşleştirilir.
    Sayfa sınırlarını aşan tarih referansları da desteklenir.
    """
    from collections import defaultdict

    def _is_all_dates(line):
        toks = line.strip().split()
        return bool(toks) and all(parse_date(t) for t in toks)

    def _is_all_times(line):
        toks = line.strip().split()
        return bool(toks) and all(re.match(r'^\d{2}:\d{2}:\d{2}$', t) for t in toks)

    def _is_num_only(line):
        toks = line.strip().split()
        if not toks: return False
        return all(re.match(r'^-?\d+([,\.]+\d+)?$', t.rstrip(',').rstrip('.'))
                   for t in toks)

    SKIP = {'ef-','ef+','Alıcının','Vericinin','Verici ','Cross','Kan Seri',
            'Yapan','Otoantikor','AÇIKLAMA','KOLONİ','BETA','HASTANE','ESBL',
            'R-PT','SAİT','ÖZTÜRK'}

    OVERFLOW_HEADERS = re.compile(
        r'[A-ZÇĞİÖŞÜ]{4,}|DİRENÇLİ|DUYARLI|ef[-+]|Alıcı|Verici|KÜLTÜR|HEMOGRAM|SAİT'
    )

    param_results = defaultdict(dict)
    ordered_params = []   # [(label, last_date)] — overflow eşleştirmesi için
    last_dates = []       # Sayfalar arası korunur

    for page_text in pages:
        lines = [l.strip() for l in page_text.split('\n') if l.strip()]
        if not lines: continue

        has_header = any(OVERFLOW_HEADERS.search(l) for l in lines)

        if not has_header:
            # OVERFLOW sayfası: başlıksız tarih+değer grupları
            i = 0
            while i < len(lines):
                if _is_all_dates(lines[i]):
                    grp_dates = [parse_date(t) for t in lines[i].split() if parse_date(t)]
                    i += 1
                    if i < len(lines) and _is_all_times(lines[i]): i += 1
                    if i < len(lines) and _is_num_only(lines[i]):
                        grp_vals = [to_float(t) for t in lines[i].split()
                                    if to_float(t) is not None]
                        # Hangi parametreye ait? Son tarihten büyük olan ilk parametreye
                        for pi, (plabel, plast_date) in enumerate(ordered_params):
                            if plast_date and all(d > plast_date for d in grp_dates):
                                for dt, v in zip(grp_dates, grp_vals):
                                    if dt and v is not None:
                                        param_results[plabel][dt] = v
                                ordered_params[pi] = (plabel, max(grp_dates))
                                break
                        i += 1
                else:
                    i += 1
            continue

        # NORMAL sayfa: sıralı satır taraması
        i = 0
        page_new_params = []

        while i < len(lines):
            line = lines[i]
            toks = line.split()

            if _is_all_dates(line):
                last_dates = [parse_date(t) for t in toks if parse_date(t)]
                i += 1; continue

            if _is_all_times(line):
                i += 1; continue

            if line.startswith('ef') or line.startswith('SAİT') or line.startswith('ÖZTÜRK'):
                i += 1; continue

            # "LABEL num num..." tespiti
            num_start = None
            for j, tok in enumerate(toks):
                ct = tok.rstrip(',').rstrip('.')
                if re.match(r'^-?\d+([,\.]+\d+)?$', ct):
                    num_start = j; break

            if num_start and num_start > 0 and last_dates:
                label = ' '.join(toks[:num_start])
                if any(label.startswith(s) or s.strip() == label.strip() for s in SKIP):
                    i += 1; continue

                # Değerleri parse et (Onaylanmadı → None)
                vals = []
                j2 = 0
                raw = toks[num_start:]
                while j2 < len(raw):
                    t = raw[j2]
                    if 'Onaylanmad' in t:
                        vals.append(None); j2 += 1
                        if j2 < len(raw) and raw[j2] == 'ı': j2 += 1
                    elif t in ('ı', '-'):
                        vals.append(None); j2 += 1
                    else:
                        vals.append(to_float(t)); j2 += 1

                for dt, v in zip(last_dates, vals):
                    if dt and v is not None:
                        param_results[label][dt] = v

                last_dt = max((d for d in last_dates if d), default=None)
                page_new_params.append((label, last_dt))

            i += 1

        if page_new_params:
            ordered_params = page_new_params

    return dict(param_results)

def lab_extract_series(full_text, param_label, pages=None, overflow_map=None, overflow_row=0):
    """Eski arayüz — geriye dönük uyumluluk için."""
    results = {}
    pattern = (
        r'((?:\d{1,2}\.\d{2}\.\d{4}\s+)+)'
        r'(?:[^\n]*\n)'
        r'(?:[^\n]*\n)'
        r'(?:ef-[^\n]*\n)?'
        + re.escape(param_label)
        + r'\s+((?:[\d,\.]+\s*)+)'
    )
    for m in re.finditer(pattern, full_text):
        dates = [parse_date(d) for d in re.findall(r'\d{1,2}\.\d{2}\.\d{4}', m.group(1))]
        vals  = [to_float(x) for x in m.group(2).strip().split()]
        for dt, v in zip(dates, vals):
            if dt and v is not None:
                results[dt] = v
    return results
# [KALDIRILDI] lab_extract_hemogram: hicbir yerden cagrilmiyor

def e1_extract_kan_gazi(full_text):
    results = {}
    blocks = re.split(r'(\d{1,2}\.\d{2}\.\d{4})\s+\d+\.GÜN', full_text)
    i = 1
    while i < len(blocks) - 1:
        dt = parse_date(blocks[i])
        icerik = re.sub(r'\s+', ' ', blocks[i+1])
        if not dt:
            i += 2; continue

        # ph, pco2, po2, be/hco3 karışık sırada
        m = re.search(
            r'[Kk]an gazı[:\s]+ph\s*([\d,\.]+)[,\s]+pco2\s*([\d,\.]+)[,\s]+'
            r'po2\s*([\d,\.]+).*?'
            r'(?:be|hco3)\s*([\-\d,\.]+)[,\s]+'
            r'(?:hco3|be)\s*([\-\d,\.]+)[,\s]+'
            r'spo2\s*([\d,\.]+)',
            icerik, re.IGNORECASE
        )
        if not m:
            # İlk gün: "ph X, pco2 X, po2 X, hco3 X, be X"
            m = re.search(
                r'ph\s*([\d,\.]+)[,\s]+pco2\s*([\d,\.]+)[,\s]+po2\s*([\d,\.]+)[,\s]+'
                r'hco3\s*([\d,\.]+)[,\s]+be\s*([\-\d,\.]+)',
                icerik, re.IGNORECASE
            )
            if m:
                results[dt] = {'ph': to_float(m.group(1)), 'pco2': to_float(m.group(2)),
                               'po2': to_float(m.group(3)), 'hco3': to_float(m.group(4)),
                               'be':  to_float(m.group(5))}
        else:
            # be/hco3 sırasını bul
            pre4 = icerik[m.start(4)-10:m.start(4)]
            if 'be' in pre4.lower():
                be_v, hco3_v = to_float(m.group(4)), to_float(m.group(5))
            else:
                hco3_v, be_v = to_float(m.group(4)), to_float(m.group(5))
            results[dt] = {'ph': to_float(m.group(1)), 'pco2': to_float(m.group(2)),
                           'po2': to_float(m.group(3)), 'hco3': hco3_v, 'be': be_v,
                           'spo2': to_float(m.group(6))}

        if dt in results:
            mn = re.search(r'sodyum\s*([\d,\.]+)', icerik, re.IGNORECASE)
            mk = re.search(r'potasyum\s*([\-\d,\.]+)', icerik, re.IGNORECASE)
            if mn: results[dt]['na'] = to_float(mn.group(1))
            if mk: results[dt]['k']  = to_float(mk.group(1))
        i += 2
    return results

def e2_extract_lab(full_text):
    norm = re.sub(r'\s+', ' ', full_text)
    results = defaultdict(dict)
    params = [
        (r'BUN \(KAN ÜRE AZOTU\)',       'bun'),
        (r'KREATİNİN \(SERUM\)',          'kreatinin'),
        (r'ÜRE',                          'ure'),
        (r'CRP \(TÜRBIDIMETRIK\)',        'crp'),
        (r'E-GFR',                        'egfr'),
        (r'HGB \(Hemoglobin\)',           'hgb'),
        (r'HCT \(Hematokrit\)',           'hct'),
        (r'LYM% \(Lenfosit Yüzdesi\)',   'lym_pct'),
        (r'LYM# \(Lenfosit Sayısı\)',    'lym_abs'),
    ]
    for pat, col in params:
        for m in re.finditer(r'(\d{1,2}\.\d{2}\.\d{4})\s+' + pat + r'[^\d\-]*([\d,\.]+)', norm):
            dt = parse_date(m.group(1))
            v  = to_float(m.group(2))
            if dt and v is not None: results[dt][col] = v
    return dict(results)
# [KALDIRILDI] detect_overflow_pages: hicbir yerden cagrilmiyor

def build_lab_dataset(lab_text, e1_text, e2_text, lab_pages=None):
    """Hibrit: LAB (evrensel parser) > E1 (kan gazı) > E2 (eksik tamamla)"""
    merged = defaultdict(dict)

    # ── 1. LAB PDF: evrensel parser (taşma otomatik) ──────────────────────
    if lab_pages:
        lab_all = parse_lab_universal(lab_pages)
        LABEL_MAP = {
            'BUN (KAN ÜRE AZOTU)':         'bun',
            'KREATİNİN (SERUM)':            'kreatinin',
            'ÜRE':                          'ure',
            'CRP (TÜRBIDIMETRIK)':          'crp',
            'E-GFR':                        'egfr',
            'PROKALSİTONİN':               'prokalsitonin',
            'PT (INR)':                     'inr',
            'WBC (Lökosit)':               'wbc',
            'HGB (Hemoglobin)':            'hgb',
            'HCT (Hematokrit)':            'hct',
            'PLT (Trombosit)':             'plt',
            'NEU% (Nötrofil Yüzdesi)':     'neu_pct',
            'CLAC':                          'laktat',
            'NEU# ( Nötrofil Sayısı)':     'neu_abs',
            'LYM# (Lenfosit Sayısı)':      'lym_abs',
            'LYM% (Lenfosit Yüzdesi)':     'lym_pct',
            'CLAC':                          'laktat',
        }
        for label, col in LABEL_MAP.items():
            for dt, v in lab_all.get(label, {}).items():
                merged[dt][col] = v

        # WBC normalize: bazı PDF'lerde 16,24 (x10³) bazılarında 16240 (/µL)
        # Kural: <100 ise x1000
        for dt in list(merged.keys()):
            if 'wbc' in merged[dt] and merged[dt]['wbc'] is not None:
                if merged[dt]['wbc'] < 100:
                    merged[dt]['wbc'] = merged[dt]['wbc'] * 1000
    else:
        # Fallback: eski yöntem
        for label, col in [
            ('BUN (KAN ÜRE AZOTU)', 'bun'),
            ('KREATİNİN (SERUM)',   'kreatinin'),
            ('ÜRE',                 'ure'),
            ('CRP (TÜRBIDIMETRIK)', 'crp'),
            ('E-GFR',               'egfr'),
            ('PROKALSİTONİN',      'prokalsitonin'),
        ]:
            for dt, v in lab_extract_series(lab_text, label).items():
                merged[dt][col] = v

    # ── 2. E1: kan gazı (birincil kaynak) ─────────────────────────────────
    for dt, vals in e1_extract_kan_gazi(e1_text).items():
        for col, v in vals.items():
            merged[dt]['kg_' + col] = v

    # ── 3. E2: eksik tarihleri tamamla veya çakışırsa ayrı kayıt ─────────
    e2_data = e2_extract_lab(e2_text)
    for dt, vals in e2_data.items():
        for col, v in vals.items():
            if col not in merged[dt]:
                # LAB'da yok → E2'den al, kaynak E2 olarak işaretle
                merged[dt][col] = v
                merged[dt]['_src_' + col] = 'E2'
            elif merged[dt][col] != v:
                # Aynı tarih, aynı parametre, farklı değer → E2 ayrı satır
                # "dt_E2" anahtarıyla saklıyoruz, write_lab'da işleneceiz
                e2_key = (dt, 'E2')
                merged[e2_key][col] = v
                merged[e2_key]['_is_e2_extra'] = True

    return dict(merged)

# ═══════════════════════════════════════════════════════════════════════════════
# MİKROBİYOLOJİ PARSER
# ═══════════════════════════════════════════════════════════════════════════════

def parse_mikrobiyoloji(paths):
    kulturler = []
    antibiyogram = []

    NUMUNE_TYPES = [
        'KAN KÜLTÜRÜ', 'TRAKEAL ASPİRAT', 'İDRAR KÜLTÜRÜ',
        'YARA KÜLTÜRÜ', 'BOS KÜLTÜRÜ', 'BAL KÜLTÜRÜ',
        'KATETER KÜLTÜRÜ', 'DOKU KÜLTÜRÜ',
    ]
    # Kötü anahtar kelimeler (yer bilgisi değil bunlar)
    # Kuruma-özgü ad tokenları herkese açık repoya konmaz; yerelde eklenir.
    BAD_YER = ['HASTANESİ', 'ÖZEL', 'Numune', 'Uzman', 'T.C.',
               'GRAM', 'KÜLTÜR', 'MİKROBİYOLOJİ', 'Referans', 'Tetkik',
               'BOYASIZ', 'BOYALI', 'AÇIKLAMA', 'ANTİBİYOGRAM',
               'DİRENÇLİ', 'DUYARLI', 'GÖRÜLDÜ', 'ÜREME', 'KOLONİ']

    for path in paths:
        text = get_pdf_text(path)
        m = re.search(r'Rapor Numarası\s*:\s*([\d\.]+)', text)
        rapor_no = m.group(1) if m else ""

        # Tarih parsing: her numune blogu icin istem ve onay tarihlerini cek
        # Format satirlari:
        #   "Aytac GUNDUZ KAN KULTURU 4.04.2022 10:21:53 4.04.2022 11:43:02"  -> istem=dates[0]
        #   "HASTANESI 4.04.2022 10:22:02 9.04.2022 12:23:00"                 -> onay=dates[1]
        istem_tarihleri = []
        onay_tarihleri = []
        lines_m = text.split('\n')
        date_pat = re.compile(r'(\d{1,2}\.\d{2}\.\d{4})\s+\d{2}:\d{2}:\d{2}')
        NUMUNE_KEYS = ['KAN KÜLTÜRÜ', 'TRAKEAL ASPİRAT', 'İDRAR KÜLTÜRÜ',
                       'YARA KÜLTÜRÜ', 'BAL KÜLTÜRÜ', 'BOS KÜLTÜRÜ',
                       'KATETER KÜLTÜRÜ', 'DOKU KÜLTÜRÜ']
        for li, line in enumerate(lines_m):
            if any(nt in line for nt in NUMUNE_KEYS):
                # Istem tarihi: bu satirda tarih varsa al
                dates_this = date_pat.findall(line)
                istem_tarihleri.append(dates_this[0] if dates_this else "")
                # Onay tarihi: sonraki 5 satirda HASTANESİ ile baslayan satirda 2. tarih
                onay_found = ""
                for next_li in range(li+1, min(li+6, len(lines_m))):
                    next_line = lines_m[next_li]
                    dates_next = date_pat.findall(next_line)
                    if len(dates_next) >= 2:
                        # Bu satirda 2 tarih var: alınma ve onay
                        onay_found = dates_next[1]
                        break
                onay_tarihleri.append(onay_found)

        # ── Numune türü + yer ──────────────────────────────────────────────
        # Rapor başlık satırı formatı:
        # "Ad Soyad   KAN KÜLTÜRÜ   15.06.2022 09:20   15.06.2022 11:20"
        # Sonraki satır: "Erişkin Yoğun Bakım"
        numune_turleri = []
        numune_yerleri = []

        for m_hdr in re.finditer(
            r'[A-ZÇĞİÖŞÜa-zçğışöü\s]+?\s+'
            r'(KAN KÜLTÜRÜ|TRAKEAL ASPİRAT|İDRAR KÜLTÜRÜ|YARA KÜLTÜRÜ'
            r'|BAL KÜLTÜRÜ|BOS KÜLTÜRÜ|KATETER KÜLTÜRÜ|DOKU KÜLTÜRÜ)'
            r'\s+\d{1,2}\.\d{2}\.\d{4}[^\n]*\n([^\n]{3,50})\n',
            text, re.IGNORECASE
        ):
            nt = m_hdr.group(1).upper().strip()
            cand = m_hdr.group(2).strip()

            # Geçerli yer mi?
            is_valid = (
                not re.search(r'\d{1,2}\.\d{2}\.\d{4}|\d{2}:\d{2}:\d{2}', cand) and
                not any(kw in cand for kw in BAD_YER) and
                re.search(r'[a-zA-ZçğışöüÇĞİŞÖÜ]{3,}', cand)
            )
            numune_turleri.append(nt)
            numune_yerleri.append(cand[:40] if is_valid else "")

        # Fallback
        if not numune_turleri:
            for nt in NUMUNE_TYPES:
                if nt in text.upper():
                    numune_turleri.append(nt)
                    numune_yerleri.append("")

        # ── Mikroorganizma listesi ─────────────────────────────────────────
        mikroorg_list = re.findall(
            r'MİKROORGANİZMA\s+([A-ZÇĞİÖŞÜa-zçğışöü\s]+?)\s+Koloni', text
        )

        for i, nt in enumerate(numune_turleri):
            tarih   = parse_date(istem_tarihleri[i]) if i < len(istem_tarihleri) else None
            s_tarih = parse_date(onay_tarihleri[i])  if i < len(onay_tarihleri)  else None
            morg    = normalize(mikroorg_list[i])     if i < len(mikroorg_list)   else ""
            yer     = numune_yerleri[i]               if i < len(numune_yerleri)  else ""

            if re.search(r'ÜREME OLMADI', text, re.IGNORECASE) and not morg:
                sonuc = "ÜREMEDI"
            elif morg:
                sonuc = "ÜREDI"
            else:
                gram = re.search(r'GRAM[^\n]+', text)
                sonuc = normalize(gram.group(0)) if gram else "BELIRSIZ"

            # Ek yer bilgisi: kültür sonuç metninde "(SOL FEMORAL)" gibi
            alim_yeri_m = re.search(r'GÖRÜLDÜ\.\(([A-ZÇĞİÖŞÜa-zçğışöü\s]+)\)', text)
            if alim_yeri_m and not yer:
                yer = alim_yeri_m.group(1).strip()

            kulturler.append({
                'rapor_no': rapor_no,
                'numune_turu': nt,
                'numune_yeri': yer,
                'numune_tarihi': tarih,
                'sonuc_tarihi': s_tarih,
                'mikroorganizma': morg,
                'sonuc': sonuc,
            })

        # ── Antibiyogram ──────────────────────────────────────────────────
        for block in re.split(r'MİKROORGANİZMA\s+', text)[1:]:
            mm = re.match(r'([A-ZÇĞİÖŞÜa-zçğışöü\s]+?)\s+Koloni', block)
            if not mm: continue
            morg_name = normalize(mm.group(1))

            def clean_abx(raw):
                for kw in ['Durum ', 'ANTİBİYOGRAM ', 'Koloni Sayısı ', 'Antibiyotik Adı ']:
                    if kw in raw: raw = raw.split(kw)[-1].strip()
                return raw

            SKIP_ABX = {"Antibiyotik Adı","ANTİBİYOGRAM","Açıklama",
                        "Tıbbi Laboratuvar Yorum","Koloni Sayısı","Durum",""}

            # Format 1: "MEROPENEM Mic: >=16 AZ DUYARLI"
            for am in re.finditer(
                r'^(?:ANTİBİYOGRAM\s+)?(.+?)\s+Mic:\s*([^\s]+)\s+(DUYARLI|DİRENÇLİ|AZ DUYARLI)',
                block, re.MULTILINE
            ):
                abx = clean_abx(normalize(am.group(1)))
                if abx in SKIP_ABX or len(abx) < 2: continue
                antibiyogram.append({
                    'rapor_no': rapor_no, 'mikroorganizma': morg_name,
                    'antibiyotik': abx, 'mic': am.group(2),
                    'duyarlilik': am.group(3),
                })

            # Format 2: "FUSİDİK ASİT DİRENÇLİ" veya "AMİKACİN (AN) DİRENÇLİ"
            for am in re.finditer(
                r'^([A-ZÇĞİÖŞÜa-zçğışöü\s\(\)\/\+\-]+?)\s+(DUYARLI|DİRENÇLİ|AZ DUYARLI)\s*$',
                block, re.MULTILINE
            ):
                abx = clean_abx(normalize(am.group(1)))
                if abx in SKIP_ABX or len(abx) < 2: continue
                antibiyogram.append({
                    'rapor_no': rapor_no, 'mikroorganizma': morg_name,
                    'antibiyotik': abx, 'mic': '',
                    'duyarlilik': am.group(2),
                })

    return kulturler, antibiyogram

def process_patient(hasta_id, pdf_map, hastane_adi="", verbose=True):
    if verbose: print(f"  [{hasta_id}] işleniyor...")
    demo = {}; e2_recs = []; lab_data = {}
    kulturler = []; antibiyogram = []; amr = []

    vazo_courses = []
    if 'E1' in pdf_map:
        try:
            demo = parse_e1(pdf_map['E1'])
            demo['protokol_no'] = hasta_id
            if hastane_adi:
                demo['hastane'] = hastane_adi
            amr = parse_antimikrobial(pdf_map['E1'])
            vazo_courses = parse_vazopressor(
                pdf_map['E1'],
                pdf_map.get('E2')
            )
            if verbose: print(f"    E1 ✓  {demo.get('ad_soyad','')} — {demo.get('yatis_tarihi','')}")
        except Exception as e:
            print(f"    [HATA] E1: {e}")

    if 'E2' in pdf_map:
        try:
            e2_recs, dogum_e2 = parse_e2(pdf_map['E2'])
            # E2'den 0. gun Apache ve beklenen olum degerini demo'ya kaydet
            if e2_recs:
                ilk = e2_recs[0]
                if not demo.get('apache_ii_0') and ilk.get('apache_ii'):
                    demo['apache_ii_0'] = ilk.get('apache_ii')
                if not demo.get('beklenen_olum_0') and ilk.get('beklenen_olum'):
                    demo['beklenen_olum_0'] = ilk.get('beklenen_olum')
            # E2'den dogum tarihi al (E1'de yoksa)
            if dogum_e2 and not demo.get('dogum_tarihi'):
                demo['dogum_tarihi'] = dogum_e2
            if verbose: print(f"    E2 ✓  {len(e2_recs)} günlük kayıt")
        except Exception as e:
            print(f"    [HATA] E2: {e}")

    if 'LAB' in pdf_map:
        try:
            # Sayfaları ayrı al (overflow tespiti için)
            with pdfplumber.open(pdf_map['LAB']) as _pdf:
                lab_pages = [p.extract_text() or "" for p in _pdf.pages]
            lab_text = "\n".join(lab_pages)
            e1_text  = get_pdf_text(pdf_map['E1']) if 'E1' in pdf_map else ""
            e2_text  = get_pdf_text(pdf_map['E2']) if 'E2' in pdf_map else ""
            lab_data = build_lab_dataset(lab_text, e1_text, e2_text, lab_pages=lab_pages)
            # Özet
            params_with_data = set()
            for d in lab_data.values():
                params_with_data.update(d.keys())
            if verbose:
                print(f"    LAB ✓  {len(lab_data)} ölçüm günü | "
                      f"parametreler: {', '.join(sorted(params_with_data))}")
        except Exception as e:
            print(f"    [HATA] LAB: {e}")

    m_paths = [pdf_map[k] for k in sorted(pdf_map) if k.startswith('M')]
    if m_paths:
        try:
            kulturler, antibiyogram = parse_mikrobiyoloji(m_paths)
            if verbose:
                print(f"    Mikro ✓  {len(kulturler)} kültür | {len(antibiyogram)} antibiyogram")
        except Exception as e:
            print(f"    [HATA] Mikrobiyoloji: {e}")

    return demo, e2_recs, lab_data, kulturler, antibiyogram, amr, vazo_courses

# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL STYLES
# ═══════════════════════════════════════════════════════════════════════════════

HDR  = PatternFill("solid", fgColor="1F4E79")
ALT  = PatternFill("solid", fgColor="D6E4F0")
HFNT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
NFNT = Font(name="Arial", size=10)
BFNT = Font(name="Arial", bold=True, size=10)
CENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left",   vertical="center", wrap_text=True)
thin = Side(style="thin", color="BFBFBF")
BRD  = Border(left=thin, right=thin, top=thin, bottom=thin)

TAB_COLORS = {
    "DEMOGRAFI":    "1F4E79",
    "KLINIK":       "2E75B6",
    "LABORATUVAR":  "70AD47",
    "KAN_GAZI":     "ED7D31",
    "ANTIMIKROBIAL":"C00000",
    "MIKROBIYOLOJI":"7030A0",
    "ANTIBIYOGRAM": "9E480E",
}

def mkhdr(ws, row, col, val, w=16):
    c = ws.cell(row=row, column=col, value=val)
    c.fill=HDR; c.font=HFNT; c.alignment=CENT; c.border=BRD
    ws.column_dimensions[get_column_letter(col)].width = w
    return c

def mkcel(ws, row, col, val, alt=False):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = ALT if alt else PatternFill()
    c.font = NFNT; c.alignment = LEFT; c.border = BRD
    return c

def setup_sheet(wb, name, freeze="A2"):
    ws = wb.create_sheet(name)
    ws.sheet_properties.tabColor = TAB_COLORS.get(name, "1F4E79")
    ws.freeze_panes = freeze
    ws.sheet_view.showGridLines = True
    return ws

# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL YAZICILAR
# ═══════════════════════════════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════════════════════════════
#  DE-IDENTIFICATION KATMANI
# ══════════════════════════════════════════════════════════════════════════════
#  AMAC: Ciktiya hicbir dogrudan tanimlayici yazilmaz. Hasta protokol numarasi
#        yalnizca PDF dosyalarini gruplamak icin BELLEKTE kullanilir; Excel'e
#        calisma-ozel takma kimlik (ICU-001, ICU-002, ...) yazilir.
#
#  MOD:
#    "standart"  -> dogrudan tanimlayicilar silinir + Hasta_ID takma kimlige
#                   cevrilir. Tarihler KORUNUR (kurum ici analiz icin gerekli).
#                   Bu dosya KURUM DISINA CIKMAZ.
#    "tam"       -> ayrica tum tarihler hasta-bazli rastgele bir gun kaydirmasi
#                   ile otelenir (hasta ICINDEKI araliklar korunur) ve dogum
#                   tarihi/yatis-cikis mutlak tarihleri silinir.
#                   Kurum disina cikacak veri setleri icin.
#
#  ANAHTAR: protokol -> takma kimlik eslesmesi AYRI bir dosyaya (key_map.csv)
#           yazilir. Bu dosya kurumda, sifreli saklanmali; PAYLASILMAMALIDIR.
#           Depoya (GitHub) ve ek dosyalara ASLA konmamalidir.
# ══════════════════════════════════════════════════════════════════════════════

import random as _random
import csv as _csv
from datetime import datetime as _dt, date as _date, timedelta as _td

DEIDENTIFY = True          # False yapilirsa ham cikti (SADECE hata ayiklama icin)
DEID_MOD = "standart"      # "standart" | "tam"
DEID_SEED = 2025           # tekrarlanabilirlik
DEID_ONEK = "ICU"

# Ciktidan TAMAMEN silinecek sutunlar (dogrudan tanimlayicilar)
_GIZLI_SUTUNLAR = ["Ad_Soyad", "TC_Kimlik", "Dogum_Tarihi", "Hastane", "Bolum", "Kurum"]
# "tam" modda ayrica silinecekler
_GIZLI_SUTUNLAR_TAM = ["Yatis_Tarihi", "Cikis_Tarihi"]

_PSEUDO = {}        # protokol_no -> "ICU-001"
_DATE_SHIFT = {}    # protokol_no -> gun kaydirmasi (int)


def _takma_kimlik(protokol):
    """Protokol numarasina kararli bir takma kimlik atar (ilk gorulme sirasina gore)."""
    p = str(protokol).strip()
    if p not in _PSEUDO:
        _PSEUDO[p] = f"{DEID_ONEK}-{len(_PSEUDO) + 1:03d}"
        if DEID_MOD == "tam":
            rng = _random.Random(f"{DEID_SEED}:{p}")
            _DATE_SHIFT[p] = rng.randint(-365, 365)   # hasta ICI araliklar korunur
    return _PSEUDO[p]


def _tarih_otele(v, protokol):
    """'tam' modda tarihi hasta-bazli sabit bir offsetle oteler."""
    if DEID_MOD != "tam" or protokol not in _DATE_SHIFT:
        return v
    off = _DATE_SHIFT[protokol]
    if isinstance(v, (_dt, _date)):
        return v + _td(days=off)
    return v


def deidentify_workbook(wb):
    """Workbook'u kaydetmeden HEMEN ONCE calistirilir.

    1) Dogrudan tanimlayici sutunlari siler (baslik adina gore, sirasindan bagimsiz)
    2) Tum sekmelerde Hasta_ID sutununu takma kimlige cevirir
    3) 'tam' modda tarihleri oteler
    Deger dondurmez; workbook yerinde degistirilir.
    """
    if not DEIDENTIFY:
        print("  !! UYARI: DEIDENTIFY=False — cikti DOGRUDAN TANIMLAYICI icerir.")
        return

    silinecek = list(_GIZLI_SUTUNLAR)
    if DEID_MOD == "tam":
        silinecek += _GIZLI_SUTUNLAR_TAM

    # Takma kimlikleri, DEMOGRAFI'deki satir sirasina gore atamak icin onceden uret
    if "DEMOGRAFI" in wb.sheetnames:
        ws0 = wb["DEMOGRAFI"]
        bas0 = {str(c.value): c.column for c in ws0[1] if c.value}
        j = bas0.get("Hasta_ID")
        if j:
            for r in range(2, ws0.max_row + 1):
                v = ws0.cell(r, j).value
                if v not in (None, ""):
                    _takma_kimlik(v)

    silinen_toplam = 0
    for ws in wb.worksheets:
        basliklar = {str(c.value): c.column for c in ws[1] if c.value}

        # --- 1) Hasta_ID -> takma kimlik  (+ 'tam' modda tarih oteleme)
        j_id = basliklar.get("Hasta_ID")
        if j_id:
            tarih_sut = [c.column for c in ws[1]
                         if c.value and ("Tarih" in str(c.value))]
            for r in range(2, ws.max_row + 1):
                ham = ws.cell(r, j_id).value
                if ham in (None, ""):
                    continue
                prot = str(ham).strip()
                if DEID_MOD == "tam":
                    for jt in tarih_sut:
                        ws.cell(r, jt).value = _tarih_otele(ws.cell(r, jt).value, prot)
                ws.cell(r, j_id).value = _takma_kimlik(prot)

        # --- 2) Gizli sutunlari sil (sagdan sola, indeks kaymasin)
        for ad in sorted(silinecek, key=lambda a: -(basliklar.get(a) or 0)):
            j = basliklar.get(ad)
            if j:
                ws.delete_cols(j)
                silinen_toplam += 1

    print(f"  De-identification: mod='{DEID_MOD}', {len(_PSEUDO)} hasta takma kimlige "
          f"cevrildi, {silinen_toplam} tanimlayici sutun silindi.")


def write_key_map(path="key_map_GIZLI.csv"):
    """Protokol -> takma kimlik anahtarini AYRI dosyaya yazar.

    !!! BU DOSYA KURUM ICINDE, SIFRELI SAKLANMALIDIR.
    !!! GitHub'a, ek dosyalara veya dergiye ASLA KONULMAZ.
    """
    if not _PSEUDO:
        return
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = _csv.writer(f)
        w.writerow(["# GIZLI — KURUM DISINA CIKARILMAZ / CONFIDENTIAL — DO NOT SHARE"])
        basliklar = ["Protokol_No", "Takma_Kimlik"]
        if DEID_MOD == "tam":
            basliklar.append("Tarih_Kaydirma_Gun")
        w.writerow(basliklar)
        for prot, tk in _PSEUDO.items():
            satir = [prot, tk]
            if DEID_MOD == "tam":
                satir.append(_DATE_SHIFT.get(prot, 0))
            w.writerow(satir)
    print(f"  !! ANAHTAR DOSYASI: {path}  — GIZLI, PAYLASILMAZ, DEPOYA KONULMAZ.")


def write_demografi(wb, patients):
    ws = setup_sheet(wb, "DEMOGRAFI", "B2")
    cols = [
        ("Hasta_ID",12),("Ad_Soyad",22),("Dogum_Tarihi",14),("Yas",7),("Cinsiyet",10),
        ("TC_Kimlik",15),("Hastane",28),("Bolum",22),("Kurum",10),
        ("Yatis_Tarihi",13),("Cikis_Tarihi",13),("Yatis_Suresi_Gun",10),
        ("Taburcu_Sekli",20),("ICD_Kodlari",40),("Tanilar",50),("Sikayet",35),
        
        ("Apache_II_Basvuru",14),("Beklenen_Olum_Basvuru",18),("Ventilasyon",16),("SVK",10),("Renal_Durum",16),("Enfeksiyon_Kaynagi",30),
        ("Beslenme_Baslangic",18),("Beslenme_Son",18),("Beslenme_Gecis_Gunu",16),
    ]
    for i,(h,w) in enumerate(cols,1): mkhdr(ws,1,i,h,w)
    for r, p in enumerate(patients, 2):
        alt = r%2==0
        row_vals = [
            p.get('protokol_no'), p.get('ad_soyad'), p.get('dogum_tarihi'),
            p.get('yas'), p.get('cinsiyet'), p.get('tc'), p.get('hastane'),
            p.get('bolum'), p.get('kurum'), p.get('yatis_tarihi'), p.get('cikis_tarihi'),
            p.get('yatis_suresi'), p.get('taburcu_sekli'), p.get('icd_kodlari'),
            p.get('tanilar'), p.get('sikayet'),
            p.get('apache_ii_0'), p.get('beklenen_olum_0'),
            p.get('ventilasyon'), p.get('svk'), p.get('renal_durum'),
            p.get('enfeksiyon_kaynagi'),
            p.get('beslenme_baslangic'), p.get('beslenme_son'), p.get('beslenme_gecis_gunu'),
        ]
        for c, v in enumerate(row_vals, 1):
            cell = mkcel(ws, r, c, v, alt)
            if c in (3,10,11) and isinstance(v, date):
                cell.number_format = "DD.MM.YYYY"
        ws.row_dimensions[r].height = 17

def zaman_noktasi(gun_no):
    if gun_no is None: return ""
    if gun_no == 0:          return "0. Gün"
    if 5  <= gun_no <= 9:    return "7. Gün (±2)"
    if 12 <= gun_no <= 16:   return "14. Gün (±2)"
    if 19 <= gun_no <= 23:   return "21. Gün (±2)"
    if 26 <= gun_no <= 30:   return "28. Gün (±2)"
    return f"Gün {gun_no}"

def write_klinik(wb, all_klinik):
    ws = setup_sheet(wb, "KLINIK", "B2")
    cols = [
        ("Hasta_ID",12),("Tarih",13),("Gun_No",8),("Zaman_Noktasi",14),
        ("YBU_Seviye",10),("APACHE_II",10),("Beklenen_Olum_pct",14),
        ("SOFA",8),("GKS_Toplam",10),
        ("Sepsis",10),("Septik_Sok",11),
    ]
    for i,(h,w) in enumerate(cols,1): mkhdr(ws,1,i,h,w)
    # Zaman noktasi hedefleri ve merkezleri
    _ZAMAN_NOKTALARI = [
        (0,  0,  0,  '0. Gün'),
        (5,  9,  7,  '7. Gün (±2)'),
        (12, 16, 14, '14. Gün (±2)'),
        (19, 23, 21, '21. Gün (±2)'),
        (26, 30, 28, '28. Gün (±2)'),
    ]

    def _sec_en_yakin(records, yatis_tarihi, son_gun):
        """Her zaman noktasi icin en yakin gunu sec, son gunu ekle."""
        secilen = {}  # {nokta_etiketi: rec}
        for rec in records:
            tarih = rec.get('tarih')
            if not tarih or not yatis_tarihi: continue
            gun_no = (tarih - yatis_tarihi).days
            # Zaman noktalarina bak
            for gmin, gmax, gmerkez, etiket in _ZAMAN_NOKTALARI:
                if gmin <= gun_no <= gmax:
                    if etiket not in secilen:
                        secilen[etiket] = (gun_no, rec)
                    else:
                        # Merkeze daha yakin mi?
                        mevcut_gun = secilen[etiket][0]
                        if abs(gun_no - gmerkez) < abs(mevcut_gun - gmerkez):
                            secilen[etiket] = (gun_no, rec)
        # Son gunu ekle (zaman noktalarindan birinde degilse)
        if son_gun is not None:
            son_etiket = f'Son Gün ({son_gun})'
            # Son gun zaten bir zaman noktasinda mi?
            son_gun_zaman = any(gmin <= son_gun <= gmax for gmin, gmax, _, _ in _ZAMAN_NOKTALARI)
            if not son_gun_zaman:
                for rec in records:
                    tarih = rec.get('tarih')
                    if not tarih or not yatis_tarihi: continue
                    if (tarih - yatis_tarihi).days == son_gun:
                        secilen[son_etiket] = (son_gun, rec)
                        break
        # Sirala ve dondur
        return [rec for _, rec in sorted(secilen.values(), key=lambda x: x[0])]

    r = 2
    for hasta_id, yatis_tarihi, records in all_klinik:
        if not records: continue
        # Son gun no hesapla
        _gun_nolar = []
        for _r in records:
            _t = _r.get('tarih')
            _gn = (_t - yatis_tarihi).days if (_t and yatis_tarihi) else None
            if _gn is not None: _gun_nolar.append(_gn)
        _son_gun = max(_gun_nolar) if _gun_nolar else None
        # En yakin gun secimi
        _secilen_records = _sec_en_yakin(records, yatis_tarihi, _son_gun)
        for rec in _secilen_records:
            tarih   = rec.get('tarih')
            gun_no  = (tarih - yatis_tarihi).days if (tarih and yatis_tarihi) else None
            if gun_no is None: continue
            alt = r%2==0
            row_vals = [
                hasta_id, tarih, gun_no, zaman_noktasi(gun_no),
                rec.get('ybu_seviye'), rec.get('apache_ii'), rec.get('beklenen_olum'),
                rec.get('sofa'), rec.get('gks_toplam'),
                rec.get('sepsis'), rec.get('septik_sok'),
            ]
            for c, v in enumerate(row_vals, 1):
                cell = mkcel(ws, r, c, v, alt)
                if c==2 and isinstance(v, date): cell.number_format="DD.MM.YYYY"
            ws.row_dimensions[r].height = 16
            r += 1

def write_lab(wb, all_lab):
    ws = setup_sheet(wb, "LABORATUVAR", "B2")
    cols = [
        ("Hasta_ID",12),("Tarih",13),("Gun_No",8),("Kaynak",10),
        ("WBC",9),("CRP",10),("Kreatinin",11),("BUN",10),
        ("AST",8),("ALT",8),("Total_Bil",10),("Direkt_Bil",10),("Albumin",10),
        ("Na",9),("HCT",9),("PLT",9),("NEU_pct",9),("Laktat",10),
    ]
    for i,(h,w) in enumerate(cols,1): mkhdr(ws,1,i,h,w)
    r = 2
    for hasta_id, yatis_tarihi, lab_dict in all_lab:
        # Normal anahtarlar: date nesneleri
        # E2 extra anahtarlar: (date, 'E2') tuple
        normal_keys = sorted([k for k in lab_dict.keys() if isinstance(k, date)])
        e2_keys     = sorted([k for k in lab_dict.keys() if isinstance(k, tuple)],
                             key=lambda x: x[0])
        all_keys = [(k, 'LAB') for k in normal_keys] + [(k[0], 'E2') for k in e2_keys]

        for tarih_info, kaynak in all_keys:
            if kaynak == 'E2':
                d = lab_dict.get((tarih_info, 'E2'), {})
            else:
                d = lab_dict.get(tarih_info, {})
            # Sadece kan gazı olmayan günleri atla (tamamen boşsa)
            non_kg = {k:v for k,v in d.items() if not k.startswith('kg_') and not k.startswith('_')}
            if not non_kg: continue
            # Sadece laktat varsa ve başka klinik değer yoksa satır açma
            # Klinik açıdan anlamlı değer var mı? (en az 1 tane olmalı)
            SKIP_ALONE = {'laktat','egfr','lym_abs','lym_pct','hgb','hct','lym_pct','inr','prokalsitonin'}
            meaningful = {k:v for k,v in non_kg.items()
                          if k not in SKIP_ALONE and v is not None}
            if not meaningful: continue
            tarih = tarih_info
            alt = r%2==0
            try:
                gun_no = (tarih - yatis_tarihi).days if (yatis_tarihi and tarih) else None
                if gun_no is not None and gun_no < 0: gun_no = None
            except: gun_no = None
            row_vals = [
                hasta_id, tarih, gun_no, kaynak,
                d.get('wbc'), d.get('crp'), d.get('kreatinin'), d.get('bun'),
                d.get('ast'), d.get('alt'), d.get('total_bil'), d.get('direkt_bil'), d.get('albumin'),
                d.get('na'), d.get('hct'), d.get('plt'), d.get('neu_pct'), d.get('laktat'),
            ]
            for c, v in enumerate(row_vals, 1):
                cell = mkcel(ws, r, c, v, alt)
                if c==2 and isinstance(v, date): cell.number_format="DD.MM.YYYY"
            ws.row_dimensions[r].height = 16
            r += 1

def write_kan_gazi(wb, all_lab):
    ws = setup_sheet(wb, "KAN_GAZI", "B2")
    cols = [
        ("Hasta_ID",12),("Tarih",13),("Gun_No",8),("Zaman_Noktasi",14),
        ("pH",8),("pCO2",9),("pO2",9),("HCO3",9),
        ("BE",8),("SpO2",9),("Na",9),("K",9),
    ]
    for i,(h,w) in enumerate(cols,1): mkhdr(ws,1,i,h,w)
    r = 2
    for hasta_id, yatis_tarihi, lab_dict in all_lab:
        normal_keys = sorted([k for k in lab_dict.keys() if isinstance(k, date)])
        e2_keys = sorted([k for k in lab_dict.keys() if isinstance(k, tuple)], key=lambda x: x[0])
        all_keys = [(k, 'LAB') for k in normal_keys] + [(k[0], 'E2') for k in e2_keys]

        for tarih_info, kaynak in all_keys:
            if kaynak == 'E2':
                d = lab_dict.get((tarih_info, 'E2'), {})
            else:
                d = lab_dict.get(tarih_info, {})
            if not any(k.startswith('kg_') for k in d):
                continue
            tarih = tarih_info
            alt = r%2==0
            try:
                gun_no = (tarih - yatis_tarihi).days if (yatis_tarihi and tarih) else None
                if gun_no is not None and gun_no < 0: gun_no = None
            except: gun_no = None
            row_vals = [
                hasta_id, tarih, gun_no, zaman_noktasi(gun_no),
                d.get('kg_ph'), d.get('kg_pco2'), d.get('kg_po2'), d.get('kg_hco3'),
                d.get('kg_be'), d.get('kg_spo2'), d.get('kg_na'), d.get('kg_k'),
            ]
            for c, v in enumerate(row_vals, 1):
                cell = mkcel(ws, r, c, v, alt)
                if c==2 and isinstance(v, date): cell.number_format="DD.MM.YYYY"
            ws.row_dimensions[r].height = 16
            r += 1
def write_antimikrobial(wb, all_amr):
    ws = setup_sheet(wb, "ANTIMIKROBIAL", "B2")
    ws.sheet_properties.tabColor = "C00000"
    cols = [
        ("Hasta_ID",12),("Ilac_Adi",20),("Etken_Madde",20),
        ("ATC",12),("AWaRe",10),("Doz",10),("Siklik",10),
        ("Uygulama_Yolu",14),("Baslangic_Tarihi",14),("Bitis_Tarihi",14),
        ("Sure_Gun",8),("Bas_Gun_No",10),("Kurs_No",8),
        ("EHU_Onayi",10),("DDD",10),("DDD_Notu",20),("TDM",8),
        ("Tedavi_Turu",18),("Uygunluk",16),("Uygunsuzluk_Nedeni",22),
        ("Deskalasyon",14),("Eskalasyon_Nedeni",22),
    ]
    MANUEL = 18  # 18. sütundan itibaren sarı
    for i,(h,w) in enumerate(cols,1):
        c = mkhdr(ws,1,i,h,w)
        if i >= MANUEL:
            c.fill = PatternFill("solid", fgColor="FFF2CC")
    ws.freeze_panes = "B2"

    from openpyxl.worksheet.datavalidation import DataValidation
    dv_t = DataValidation(type="list", formula1='"Ampirik,Etkene özgü,Profilaktik,Eskalasyon,Deskalasyon"', showDropDown=False)
    dv_u = DataValidation(type="list", formula1='"Uygun,Uygun değil,Değerlendirilemez"', showDropDown=False)
    dv_n = DataValidation(type="list", formula1='"Yanlış ilaç,Yanlış doz,Geniş spektrum,Gereksiz,Süre uzun,Süre kısa,—"', showDropDown=False)
    dv_d = DataValidation(type="list", formula1='"Yapıldı,Yapılmadı,Gerekmedi"', showDropDown=False)
    for dv in [dv_t, dv_u, dv_n, dv_d]: ws.add_data_validation(dv)

    aware_col = {"Access":"E2EFDA","Watch":"FFF2CC","Reserve":"FCE4D6"}
    r = 2
    for hasta_id, courses in all_amr:
        for cd in sorted(courses, key=lambda x: x.get("baslangic") or date.min):
            aware = cd.get("aware","")
            vals = [
                hasta_id, cd.get("ilac_adi",""), cd.get("etken_madde",""),
                cd.get("atc",""), aware, cd.get("doz",""), cd.get("siklik",""),
                cd.get("uygulama_yolu",""), cd.get("baslangic"), cd.get("bitis"),
                cd.get("sure_gun"), cd.get("baslangic_gun_no"), cd.get("kurs_no"),
                cd.get("ehu_onayi",""), cd.get("ddd"), cd.get("ddd_notu",""),
                cd.get("tdm",""),
                cd.get("tedavi_turu",""), cd.get("uygunluk",""),
                cd.get("uygunsuzluk_nedeni",""), cd.get("deskalasyon",""),
                cd.get("eskalasyon_nedeni",""),
            ]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=r, column=ci, value=v)
                cell.font = NFNT; cell.alignment = LEFT; cell.border = BRD
                if ci == 5 and aware in aware_col:
                    cell.fill = PatternFill("solid", fgColor=aware_col[aware])
                    cell.font = BFNT
                elif ci in (9,10) and isinstance(v, date):
                    cell.number_format = "DD.MM.YYYY"
                elif ci >= MANUEL:
                    cell.fill = PatternFill("solid", fgColor="FFFDE7")
                else:
                    cell.fill = PatternFill()

            # Dropdown bağla
            cols_map = {18: dv_t, 19: dv_u, 20: dv_n, 21: dv_d}
            for ci, dv in cols_map.items():
                col_letter = ws.cell(row=r, column=ci).column_letter
                dv.add(f"{col_letter}{r}")

            ws.row_dimensions[r].height = 18
            r += 1

    # Not satırı
    nc = ws.cell(row=r+1, column=MANUEL, value="★ Sarı sütunlar manuel girilecek (dropdown seçim)")
    nc.font = Font(name="Arial", italic=True, size=9, color="7F6000")


def write_vazopressor(wb, all_vazo):
    ws = setup_sheet(wb, "VAZOPRESSOR", "B2")
    ws.sheet_properties.tabColor = "FF0000"
    cols = [
        ("Hasta_ID",12), ("Ilac",16),
        ("Kurs_No",10), ("Baslangic_Tarihi",16), ("Bitis_Tarihi",16),
        ("Sure_Gun",10), ("Baslangic_Gun_No",16), ("Bitis_Gun_No",14),
    ]
    for i,(h,w) in enumerate(cols,1): mkhdr(ws,1,i,h,w)
    ws.freeze_panes = "B2"

    # İlaç renkleri
    ilac_colors = {
        "Noradrenalin": "FCE4D6",
        "Adrenalin":    "EAD1DC",
        "Dopamin":      "E2EFDA",
        "Dobutamin":    "FFF2CC",
    }

    r = 2
    for hasta_id, yatis_tarihi, courses in all_vazo:
        # İlaç adına göre sırala, sonra başlangıç tarihine göre
        for course in sorted(courses, key=lambda x: (x['ilac'], x['baslangic'])):
            ilac    = course['ilac']
            bas     = course['baslangic']
            bit     = course['bitis']
            sure    = course['sure_gun']
            kurs_no = course['kurs_no']
            try:
                bas_gun = (bas - yatis_tarihi).days if (bas and yatis_tarihi) else None
                bit_gun = (bit - yatis_tarihi).days if (bit and yatis_tarihi) else None
            except:
                bas_gun = bit_gun = None

            row_vals = [hasta_id, ilac, kurs_no, bas, bit, sure, bas_gun, bit_gun]
            color = ilac_colors.get(ilac, "FFFFFF")
            for c, v in enumerate(row_vals, 1):
                cell = mkcel(ws, r, c, v)
                cell.fill = PatternFill("solid", fgColor=color)
                if c in (4,5) and isinstance(v, date):
                    cell.number_format = "DD.MM.YYYY"
            ws.row_dimensions[r].height = 17
            r += 1


def write_mikrobiyoloji_abg(wb, all_kultur, all_abg):
    """Mikrobiyoloji + antibiyogram birleşik sheet."""
    ws = setup_sheet(wb, "MIKROBIYOLOJI_ABG", "B2")
    ws.sheet_properties.tabColor = "7030A0"
    cols = [
        ("Hasta_ID",12),("Rapor_No",24),("Numune_Turu",20),("Numune_Yeri",22),
        ("Numune_Tarihi",14),("Sonuc_Tarihi",14),
        ("Mikroorganizma",28),("Kultur_Sonucu",16),
        ("Antibiyotik",28),("MIC",10),("Duyarlilik",14),
    ]
    for i,(h,w) in enumerate(cols,1): mkhdr(ws,1,i,h,w)
    duy_colors = {"DUYARLI":"E2EFDA","AZ DUYARLI":"FFF2CC","DİRENÇLİ":"FCE4D6"}
    KULTUR_FILL = PatternFill("solid", fgColor="EBF3FB")
    r = 2
    for (hk, kulturler),(ha, abg_list) in zip(all_kultur, all_abg):
        abg_idx = {}
        for a in abg_list:
            key = (a.get("rapor_no",""), a.get("mikroorganizma",""))
            abg_idx.setdefault(key,[]).append(a)
        for k in kulturler:
            rapor = k.get("rapor_no",""); morg = k.get("mikroorganizma",""); sonuc = k.get("sonuc","")
            abg_rows = abg_idx.get((rapor, morg), [])
            if not abg_rows or sonuc == "ÜREMEDI":
                row_vals = [hk, rapor, k.get("numune_turu"), k.get("numune_yeri",""),
                            k.get("numune_tarihi"), k.get("sonuc_tarihi"),
                            morg if morg else "—", sonuc, "", "", ""]
                for ci,v in enumerate(row_vals,1):
                    cell = mkcel(ws,r,ci,v); cell.fill = KULTUR_FILL
                    if ci in (4,5) and isinstance(v,date): cell.number_format="DD.MM.YYYY"
                ws.row_dimensions[r].height=17; r+=1
            else:
                for a in abg_rows:
                    duy = a.get("duyarlilik","")
                    row_vals = [hk, rapor, k.get("numune_turu"), k.get("numune_yeri",""),
                                k.get("numune_tarihi"), k.get("sonuc_tarihi"),
                                morg, sonuc, a.get("antibiyotik",""), a.get("mic",""), duy]
                    for ci,v in enumerate(row_vals,1):
                        cell = mkcel(ws,r,ci,v)
                        if ci<=8: cell.fill = KULTUR_FILL
                        if ci==10 and duy in duy_colors:
                            cell.fill=PatternFill("solid",fgColor=duy_colors[duy]); cell.font=BFNT
                        if ci in (4,5) and isinstance(v,date): cell.number_format="DD.MM.YYYY"
                    ws.row_dimensions[r].height=17; r+=1
# [KALDIRILDI] write_antibiyogram: hicbir yerden cagrilmiyor

# ═══════════════════════════════════════════════════════════════════════════════
# ANA PIPELINE
# ═══════════════════════════════════════════════════════════════════════════════

def run(klasor, cikti, verbose=True):
    print(f"\n{'='*60}")
    print(f"  YBÜ PDF → Excel Pipeline  v6.0")
    print(f"  Klasör : {klasor}")
    print(f"  Çıktı  : {cikti}")
    print(f"{'='*60}\n")

    pdf_groups = group_pdfs(klasor)
    if not pdf_groups:
        print("Hiç hasta bulunamadı. Dosya adlandırma kuralını kontrol edin.")
        return

    hastane_adi = get_hastane_adi(klasor)
    if hastane_adi:
        print(f"Hastane: {hastane_adi}")
    print(f"{len(pdf_groups)} hasta bulundu.\n")

    all_demo=[]; all_klinik=[]; all_lab=[]
    all_amr=[]; all_kultur=[]; all_abg=[]; all_vazo=[]

    sureler = []          # [v6] (hasta_id, pdf_sayisi, saniye)
    t_kohort0 = time.perf_counter()

    for hasta_id, pdf_map in sorted(pdf_groups.items()):
        t0 = time.perf_counter()                                  # [v6]
        demo, e2_recs, lab_dict, kulturler, abg, amr, vazo = \
            process_patient(hasta_id, pdf_map, hastane_adi=hastane_adi, verbose=verbose)
        dt = time.perf_counter() - t0                             # [v6]
        sureler.append((hasta_id, len(pdf_map), dt))              # [v6]

        yatis = demo.get('yatis_tarihi')
        all_demo.append(demo)
        all_klinik.append((hasta_id, yatis, e2_recs))
        all_lab.append((hasta_id, yatis, lab_dict))
        all_amr.append((hasta_id, amr))
        all_kultur.append((hasta_id, kulturler))
        all_abg.append((hasta_id, abg))
        all_vazo.append((hasta_id, yatis, vazo))

    print(f"\nExcel yazılıyor → {cikti}")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    write_demografi(wb, all_demo)
    write_klinik(wb, all_klinik)
    write_lab(wb, all_lab)
    write_kan_gazi(wb, all_lab)
    write_antimikrobial(wb, all_amr)
    write_vazopressor(wb, all_vazo)
    write_mikrobiyoloji_abg(wb, all_kultur, all_abg)

    # ---- De-identification: kaydetmeden ONCE ----

    deidentify_workbook(wb)

    wb.save(cikti)

    write_key_map()


    # Özet
    total_lab_rows = sum(
        sum(1 for d in ld.values() if any(not k.startswith('kg_') for k in d))
        for _,_,ld in all_lab
    )
    total_kg_rows  = sum(
        sum(1 for d in ld.values() if any(k.startswith('kg_') for k in d))
        for _,_,ld in all_lab
    )
    total_kul = sum(len(k) for _,k in all_kultur)
    total_abg = sum(len(a) for _,a in all_abg)

    # ---- [v6] Hasta basina sure logu ----------------------------------
    t_kohort = time.perf_counter() - t_kohort0
    sure_csv = str(Path(cikti).with_name(Path(cikti).stem + "_sureler.csv"))
    with open(sure_csv, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        # [v6] DE-ID: ham protokol numarasi YAZILMAZ. deidentify_workbook()
        # bu noktada calismis oldugu icin _PSEUDO doludur; takma kimlik kullanilir.
        # Boylece sure CSV'si makale ekine dogrudan konulabilir.
        w.writerow(["Hasta_ID", "PDF_sayisi", "Sure_saniye"])
        for hid, npdf, dt in sureler:
            w.writerow([_PSEUDO.get(str(hid).strip(), "?"), npdf, f"{dt:.3f}"])

    _s = sorted(d for _, _, d in sureler)
    if _s:
        _med = statistics.median(_s)
        _ort = statistics.fmean(_s)
        if len(_s) >= 4:
            _q1, _, _q3 = statistics.quantiles(_s, n=4)
        else:
            _q1 = _q3 = _med
        print(f"\n{'─'*60}")
        print(f"  SÜRE  (Bölüm 3.4 / Şekil 3 için)")
        print(f"  Hasta sayısı          : {len(_s)}")
        print(f"  Toplam çalışma süresi : {t_kohort:.1f} sn  ({t_kohort/60:.1f} dk)")
        print(f"  Hasta başına  medyan  : {_med:.2f} sn")
        print(f"                ortalama: {_ort:.2f} sn")
        print(f"                IQR     : {_q1:.2f} – {_q3:.2f} sn")
        print(f"                aralık  : {min(_s):.2f} – {max(_s):.2f} sn")
        print(f"  Hasta başına süreler  → {sure_csv}")
        # Manuel cikarim medyani 26 dk/hasta (bkz. Yontem 2.5)
        print(f"  KIYAS: manuel medyan 26 dk/hasta  →  hız kazancı ~{(26*60)/_med:.0f}×")
    # -------------------------------------------------------------------

    print(f"\n{'─'*60}")
    print(f"  ✓ {len(all_demo)} hasta işlendi")
    print(f"  ✓ Laboratuvar: {total_lab_rows} satır")
    print(f"  ✓ Kan gazı   : {total_kg_rows} satır")
    print(f"  ✓ Kültür     : {total_kul} sonuç")
    print(f"  ✓ Antibiyogram: {total_abg} satır")
    print(f"  → {cikti}")
    print(f"{'─'*60}\n")

if __name__ == "__main__":
        parser = argparse.ArgumentParser(description="YBÜ PDF → Excel (offline)")
        parser.add_argument("--klasor", default=".", help="PDF klasörü")
        parser.add_argument("--cikti",  default="ICU_VeriTabani.xlsx", help="Çıktı dosyası")
        parser.add_argument("--sessiz", action="store_true")
        args = parser.parse_args()
        run(args.klasor, args.cikti, verbose=not args.sessiz)
